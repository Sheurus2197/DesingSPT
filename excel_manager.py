#Conteene la clase ExcelManager para la gestión de archivos Excel.

from openpyxl import load_workbook, Workbook
from tkinter import filedialog, messagebox


class ExcelManager:
    """
    Clase para gestionar las operaciones con archivos Excel.
    """

    def __init__(self):
        self.file_path = None
        self.workbook = None
        self.worksheet = None

    def archivo_cargado(self):
        """
        Verifica si un archivo Excel está cargado.
        """
        return self.workbook is not None and self.file_path is not None

    def abrir_archivo(self):
        """
        Abre un archivo Excel y establece el workbook y worksheet activos.
        """
        file_path = filedialog.askopenfilename(
            title="Abrir archivo",
            filetypes=(("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*"))
        )
        if file_path:
            try:
                self.file_path = file_path
                self.workbook = load_workbook(file_path)
                self.worksheet = self.workbook.active
                # Validar que sea un archivo válido
                if self.leer_dato("A1") != "DesignSPT_DAPR_ADBD":
                    messagebox.showerror("Error", "Archivo incompatible. Código inválido en la celda A1.")
                    self.file_path, self.workbook, self.worksheet = None, None, None
                    return False
                return True
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo abrir el archivo: {e}")
        return False

    def guardar_archivo(self):
        """
        Guarda el archivo Excel actualmente abierto.
        """
        if self.archivo_cargado():
            try:
                self.workbook.save(self.file_path)
                messagebox.showinfo("Éxito", "Archivo guardado correctamente.")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar el archivo: {e}")
        else:
            messagebox.showwarning("Advertencia", "No hay ningún archivo cargado para guardar.")

    def leer_dato(self, celda):
        """
        Lee un dato de una celda específica.
        """
        if self.archivo_cargado():
            try:
                return self.worksheet[celda].value
            except Exception as e:
                print(f"Error al leer celda {celda}: {e}")
        return None

    def escribir_dato(self, celda, valor):
        """
        Escribe un dato en una celda específica.
        """
        if self.archivo_cargado():
            try:
                self.worksheet[celda] = valor
            except Exception as e:
                print(f"Error al escribir en celda {celda}: {e}")

    def crear_nuevo_archivo(self):
        """
        Crea un nuevo archivo Excel con un formato básico.
        """
        file_path = filedialog.asksaveasfilename(
            title="Crear Nuevo Archivo Excel",
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")]
        )
        if file_path:
            try:
                self.file_path = file_path
                self.workbook = Workbook()
                self.worksheet = self.workbook.active
                self.worksheet["A1"] = "DesignSPT_DAPR_ADBD"
                self.workbook.save(file_path)
                messagebox.showinfo("Éxito", "Archivo creado correctamente.")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo crear el archivo: {e}")
