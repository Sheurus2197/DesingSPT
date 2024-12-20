import tkinter as tk
from tkinter import Menu, filedialog, messagebox, ttk
from numpy.f2py.crackfortran import endifs
from tkcalendar import DateEntry
import os
from openpyxl import load_workbook, Workbook
from datetime import datetime
import math
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
import ctypes
from openpyxl.drawing.image import Image  # Para insertar imágenes en Excel


myappid = 'mycompay.myproduct.subproduct.version'
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)


# Crear la ventana principal
root = tk.Tk()
root.title("DesingSPT")

# Obtener las dimensiones de la pantalla
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()-80

# Establecer la geometría de la ventana al tamaño de la pantalla
root.geometry(f"{screen_width}x{screen_height}")

root.iconbitmap('icon.ico')

# Crear un Frame principal para organizar en columnas
main_frame = tk.Frame(root)
main_frame.pack(fill="both", expand=True)

# Frame para la primera columna
Columna_1 = tk.Frame(main_frame)
Columna_1.grid(row=0, column=0,padx=5, pady=5)

# Frame para la segunda columna
Columna_2 = tk.Frame(main_frame)
Columna_2.grid(row=0, column=1, padx=5, pady=5)

# Frame para la segunda columna (resistividad_canvas)
Columna_3 = tk.Frame(main_frame)
Columna_3.grid(row=0, column=2, padx=5, pady=5)

#Frame de proyecto
Infor_proyecto_frame = tk.LabelFrame(Columna_1, text="Información del Proyecto",pady=5,padx=5)
Infor_proyecto_frame.grid(row=0, column=0, padx=5, pady=5, rowspan=2)
tk.Label(Infor_proyecto_frame, text="Título del proyecto: ").grid(row=0, column=0, sticky="e")
tk.Label(Infor_proyecto_frame, text="Tensión Primario: ").grid(row=1, column=0, sticky="e")
tk.Label(Infor_proyecto_frame, text="Tensión Secundario: ").grid(row=2, column=0, sticky="e")
tk.Label(Infor_proyecto_frame, text="Potencia Nominal:").grid(row=3, column=0, sticky="e")

titulo_proyecto_val = tk.Label(Infor_proyecto_frame, text="")
tension_primario_val = tk.Label(Infor_proyecto_frame, text="")
tension_secundario_val = tk.Label(Infor_proyecto_frame, text="")
potencia_nominal_val = tk.Label(Infor_proyecto_frame, text="")

titulo_proyecto_val.grid(row=0, column=1, sticky="w")
tension_primario_val.grid(row=1, column=1, sticky="w")
tension_secundario_val.grid(row=2, column=1, sticky="w")
potencia_nominal_val.grid(row=3, column=1, sticky="w")

Objetivo_frame = tk.LabelFrame(Columna_1, text="Objetivos",pady=5,padx=5)
Objetivo_frame.grid(row=2, column=0, padx=5, pady=5, rowspan=5)

canvas_2d_3d = tk.Frame(Columna_2,pady=5,padx=5)
canvas_2d_3d.grid(row=0, column=0, padx=5, pady=5, rowspan=7, columnspan=2)

Resistencia_frame = tk.LabelFrame(Columna_3, text="Resistencia",pady=5,padx=5)
Resistencia_frame.grid(row=0, column=0, padx=5, pady=5, rowspan=2)

Seguridad_frame = tk.LabelFrame(Columna_3, text="Seguridad",pady=5,padx=5)
Seguridad_frame.grid(row=2, column=0, padx=5, pady=5, rowspan=2)

Default_frame = tk.LabelFrame(Columna_3, text="Default",pady=5,padx=5)
Default_frame.grid(row=4, column=0, padx=5, pady=5, rowspan=2)

boton_Calcular_frame = tk.Frame(Columna_3, pady=5, padx=5)
boton_Calcular_frame.grid(row=6, column=0, padx=5, pady=5)



class ExcelManager:
    def __init__(self):
        self.file_path = None
        self.workbook = None
        self.worksheet = None

    def archivo_cargado(self):
        """Verifica si un archivo Excel está cargado."""
        return self.workbook is not None and self.file_path is not None

    def abrir_archivo(self):
        """Abre un archivo Excel y establece el workbook y worksheet activos."""
        file_path = filedialog.askopenfilename(
            title="Abrir archivo",
            filetypes=(("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*"))
        )
        if file_path:
            try:
                self.file_path = file_path
                self.workbook = load_workbook(file_path)
                self.worksheet = self.workbook.active
                return True
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo abrir el archivo: {e}")
        return False

    def guardar_archivo(self):
        """Guarda el archivo Excel actualmente abierto."""
        if self.file_path and self.workbook:
            try:
                self.workbook.save(self.file_path)
                messagebox.showinfo("Guardado", "El archivo se ha guardado correctamente.")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar el archivo: {e}")
        else:
            messagebox.showwarning("Advertencia", "No hay ningún archivo abierto para guardar.")

    def leer_dato(self, celda):
        """Lee un dato de una celda específica."""
        if self.worksheet:
            return self.worksheet[celda].value
        return None

    def escribir_dato(self, celda, valor):
        """Escribe un dato en una celda específica."""
        if self.worksheet:
            self.worksheet[celda] = valor

    def crear_nuevo_archivo(self):
        """Crea un nuevo archivo Excel con un formato básico."""
        try:
            file_path = filedialog.asksaveasfilename(
                title="Crear Nuevo Archivo Excel",
                defaultextension=".xlsx",
                filetypes=[("Archivos Excel", "*.xlsx")]
            )
            if not file_path:
                return

            # Crear un nuevo archivo Excel
            self.workbook = Workbook()
            self.file_path = file_path
            self.worksheet = self.workbook.active
            self.worksheet["A1"] = "DesignSPT_DAPR_ADBD"
            self.workbook.save(file_path)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo crear el archivo: {e}")
            return False


manager = ExcelManager()


# Función para abrir un archivo
def abrir_archivo():
    """
        Función para seleccionar y abrir un archivo, estableciendo la ruta y validando su contenido.
        Si el contenido de la celda A1 no es 'DesignSPT_DAPR_ADBD', se mostrará un error y no se abrirá.
        """
    if manager.abrir_archivo():
        try:
            # Leer el contenido de la celda A1
            codigo_validacion = manager.leer_dato("A1")

            # Validar el código de identificación
            if codigo_validacion != "DesignSPT_DAPR_ADBD":
                messagebox.showerror("Error de archivo", "El archivo no es compatible. Código inválido en A1.")
                manager.file_path = None  # Limpiar el archivo cargado
                manager.workbook = None
                manager.worksheet = None
                return

            # Habilitar las opciones del menú si el archivo es válido
            habilitar_menu_edit()
            actualizar_datos_principal()


        except Exception as e:
            messagebox.showerror("Error", f"Hubo un problema al validar el archivo: {e}")
            manager.file_path = None  # Limpiar el archivo cargado
            manager.workbook = None
            manager.worksheet = None


def guardar_archivo():
    manager.guardar_archivo()


def actualizar_datos_principal():
    """
    Obtiene datos clave de las hojas 'Información' y 'Transformador'.
    Muestra los valores en consola.
    """
    try:
        if not manager.archivo_cargado():
            print("No hay ningún archivo cargado.")
            return

        # Obtener el título del proyecto desde la hoja 'Información'
        if "Información" in manager.workbook.sheetnames:
            info_sheet = manager.workbook["Información"]
            titulo_proyecto = info_sheet.cell(row=2, column=1).value  # Celda A1
        else:
            titulo_proyecto = "No disponible"
        print(f"Título del Proyecto: {titulo_proyecto}")

        # Obtener tensión nominal y potencia nominal desde la hoja 'Transformador'
        if "trafos" in manager.workbook.sheetnames:
            transformador_sheet = manager.workbook["trafos"]
            tension_primario = transformador_sheet.cell(row=2, column=1).value  # Celda A1
            tension_secundario = transformador_sheet.cell(row=2, column=1).value  # Celda A2
            potencia_nominal = transformador_sheet.cell(row=2, column=3).value  # Celda A12
        else:
            tension_primario = "No disponible"
            tension_secundario = "No disponible"
            potencia_nominal = "No disponible"

        print(f"Tensión Primario: {tension_primario}")
        print(f"Tensión Secundario: {tension_secundario}")
        print(f"Potencia Nominal: {potencia_nominal}")

    except Exception as e:
        print(f"Error al obtener datos del proyecto: {e}")

    root.title(f"DesingSPT - {os.path.basename(manager.file_path)}")


def abrir_nuevo():
    manager.crear_nuevo_archivo()


    # Habilitar las opciones del menú si el archivo es válido
    habilitar_menu_edit()
    actualizar_datos_principal()


# Función para abrir la ventana de "Área para el SPT"

def abrir_ventana_spt():
    """
    Abre la ventana de "Geometría Definida" para seleccionar y dibujar geometrías.
    Si existen datos en la hoja "Área" del archivo Excel, los muestra en el canvas.
    """
    ventana_geometria = tk.Toplevel(root)
    ventana_geometria.title("Geometría Definida")
    ventana_geometria.geometry("500x500")

    # Canvas para dibujo
    canvas = tk.Canvas(ventana_geometria, width=400, height=400, bg="white")
    canvas.pack(pady=10)

    # Frame para controles
    control_frame = tk.Frame(ventana_geometria)
    control_frame.pack(pady=10)

    # Listbox para seleccionar geometría
    tk.Label(control_frame, text="Geometría").grid(row=0, column=0, padx=5)
    geometry_var = tk.StringVar(value="L")
    geometry_combo = ttk.Combobox(
        control_frame, textvariable=geometry_var,
        values=["L", "Rectángulo", "Línea"]
    )
    geometry_combo.grid(row=0, column=1, padx=5)

    # Botones
    tk.Button(control_frame, text="Editar", command=lambda: editar_geometria()).grid(row=0, column=2, padx=5)
    tk.Button(control_frame, text="Guardar", command=lambda: guardar_geometria()).grid(row=0, column=3, padx=5)

    def calcular_longitudes_y_perimetro(puntos, tipo_geometria):
        longitudes = []
        perimetro = 0
        for i in range(len(puntos) - (1 if tipo_geometria in ["L", "Línea"] else 0)):
            x1, y1 = puntos[i]
            x2, y2 = puntos[(i + 1) % len(puntos)]
            distancia = round(((x2 - x1) ** 2 + (y2 - y1) ** 2) ** 0.5, 2)
            longitudes.append(distancia)
            perimetro += distancia
        return longitudes, perimetro

    def calcular_dimensiones_geometria(puntos, tipo_geometria):
        """
        Calcula las longitudes de los segmentos y el perímetro de la figura según su tipo.

        :param puntos: Lista de puntos [(x1, y1), (x2, y2), ...] que definen la figura.
        :param tipo_geometria: Tipo de geometría ('L', 'Rectángulo', 'Triángulo', 'Línea', 'Circunferencia').
        :return: Una tupla con los valores específicos (dimensiones) y el perímetro.
        """
        longitudes = []
        perimetro = 0

        # Calcular las longitudes entre los puntos
        for i in range(len(puntos) - (1 if tipo_geometria in ["L", "Línea"] else 0)):
            x1, y1 = puntos[i]
            x2, y2 = puntos[(i + 1) % len(puntos)]  # Conexión cíclica para figuras cerradas
            distancia = round(((x2 - x1) ** 2 + (y2 - y1) ** 2) ** 0.5, 2)
            longitudes.append(distancia)
            perimetro += distancia

        # Devolver valores específicos según el tipo de geometría
        if tipo_geometria == "L":
            # Vertical y Horizontal
            segmento_vertical = longitudes[0] if len(longitudes) > 0 else 0
            segmento_horizontal = longitudes[1] if len(longitudes) > 1 else 0
            valores = [segmento_vertical, segmento_horizontal]
        elif tipo_geometria == "Rectángulo":
            # Base y Altura (primeros dos lados)
            base = longitudes[0] if len(longitudes) > 0 else 0
            altura = longitudes[1] if len(longitudes) > 1 else 0
            valores = [base, altura]
        elif tipo_geometria == "Triángulo":
            # Tres lados
            valores = longitudes[:3]  # Asegurar solo tres lados
        elif tipo_geometria == "Línea":
            # Longitud de la línea
            valores = [longitudes[0]] if len(longitudes) > 0 else [0]
        elif tipo_geometria == "Circunferencia":
            # Calcular el diámetro si hay dos puntos (suponiendo que son el diámetro)
            if len(puntos) == 2:
                x1, y1 = puntos[0]
                x2, y2 = puntos[1]
                diametro = round(((x2 - x1) ** 2 + (y2 - y1) ** 2) ** 0.5, 2)
            else:
                diametro = 0
            valores = [diametro]
        else:
            # Tipo no reconocido
            valores = []

        return valores

    def obtener_dimensiones_predeterminadas(tipo_geometria):
        dimensiones = {
            "L": [150, 50],
            "Rectángulo": [100, 150],
            "Triángulo": [100, 100,100],
            "Línea": [150],
            "Circunferencia": [100],
        }
        return dimensiones.get(tipo_geometria, [])

    def validar_puntos(puntos):
        if not isinstance(puntos, list) or not all(isinstance(p, tuple) and len(p) == 2 for p in puntos):
            raise ValueError("Los puntos no tienen el formato esperado.")

    figura_desde_excel = False
    # Función para manejar el evento de cambio en el combobox
    def on_geometry_change(event):
        global figura_desde_excel
        nueva_geometria = geometry_var.get()

        if figura_desde_excel:  # Solo preguntar si la figura fue dibujada desde datos del Excel
            respuesta = messagebox.askyesno(
                "Confirmar Cambio",
                "La figura actual fue cargada desde el Excel. ¿Está seguro de cambiar la geometría?"
            )
            if respuesta:  # Si el usuario confirma, dibujar con dimensiones predeterminadas
                dimensiones = obtener_dimensiones_predeterminadas(nueva_geometria)
                dibujar_figura(canvas, nueva_geometria, dimensiones)
                figura_desde_excel = False  # Desactivar la funcionalidad para futuras selecciones
        else:
            # Si no fue dibujada desde Excel, cambiar directamente
            geometry_var.set(canvas.geometry_actual)

    # Función para dibujar geometrías
    def draw_geometry():
        global figura_desde_excel
        if manager.archivo_cargado() and "Área" in manager.workbook.sheetnames:
            try:
                area_sheet = manager.workbook["Área"]
                if area_sheet.max_row >= 2:
                    fila_datos = 2
                    tipo_geometria = area_sheet.cell(row=fila_datos, column=1).value
                    puntos_str = area_sheet.cell(row=fila_datos, column=2).value
                    print("puntos_str1", puntos_str)
                    if puntos_str:
                        puntos = eval(puntos_str)
                        print("puntos1", puntos)
                        validar_puntos(puntos)
                        dimensiones = calcular_dimensiones_geometria(puntos, tipo_geometria)
                        print("puntos_str2", dimensiones)
                        geometry_var.set(tipo_geometria)
                        dibujar_figura(canvas, tipo_geometria, dimensiones)
                        figura_desde_excel = True  # Marcar como dibujada desde datos del Excel

                        print("puntos_str3",dimensiones)
                        return
            except Exception as e:
                messagebox.showerror("Error", f"No se pudieron cargar los datos del área: {e}")

        # Dibujar con dimensiones predeterminadas
        geometry_type = geometry_var.get()
        dimensiones = obtener_dimensiones_predeterminadas(geometry_type)
        dibujar_figura(canvas, geometry_type, dimensiones)
        figura_desde_excel = False  # Figura no fue dibujada desde datos del Excel

        print("draw_geometry", dimensiones)


    def dibujar_figura(canvas, tipo, dimensiones):
        """
        Dibuja una figura en el canvas según el tipo y las dimensiones dadas.

        :param canvas: Canvas donde se dibuja la figura.
        :param tipo: Tipo de figura ('L', 'Rectángulo', 'Triángulo', 'Línea', 'Circunferencia').
        :param dimensiones: Lista de dimensiones necesarias para la figura.
        """
        global points_g  # Declarar points_g como variable global
        canvas.delete("all")  # Limpiar el canvas antes de dibujar
        print("Dibujar figura - dimensiones:", dimensiones)



        canvas_width, canvas_height = int(canvas["width"]), int(canvas["height"])

        if tipo == "L":
            vertical, horizontal = dimensiones
            points = [(50, 50), (50, 50 + vertical), (50 + horizontal, 50 + vertical)]
        elif tipo == "Rectángulo":
            base, altura = dimensiones
            points = [(50, 50), (50 + base, 50), (50 + base, 50 + altura), (50, 50 + altura)]
        elif tipo == "Triángulo":
            s1, s2, s3 = dimensiones  # Puedes expandir para calcular puntos del triángulo según los lados
            points = [(50, 50), (50 + s1, 50), (50, 50 + s2)]
        elif tipo == "Línea":
            longitud = dimensiones[0]
            points = [(50, 100), (50 + longitud, 100)]
        elif tipo == "Circunferencia":
            diametro = dimensiones[0]
            radius = diametro / 2
            center_x, center_y = canvas_width // 2, canvas_height // 2
            # Dibujar la circunferencia
            canvas.create_oval(
                center_x - radius, center_y - radius,
                center_x + radius, center_y + radius,
                outline="blue", width=2
            )
            # Dibujar el diámetro
            canvas.create_line(
                center_x - radius, center_y,
                center_x + radius, center_y,
                fill="red", width=2
            )
            canvas.create_text(
                center_x, center_y - 10,
                text=f"{diametro} cm",
                font=("Arial", 8), fill="black"
            )
            perimeter = round(2 * math.pi * radius, 1)
            canvas.create_text(
                center_x, center_y + radius + 10,
                text=f"{perimeter} cm",
                font=("Arial", 8), fill="black"
            )
            return  # Termina aquí para círculos, ya que no necesitan más procesamiento

        # Determinar escalado si los puntos exceden el tamaño del canvas
        max_x = max(p[0] for p in points)
        max_y = max(p[1] for p in points)
        scale_factor = min(canvas_width / max_x if max_x > canvas_width else 1,
                           canvas_height / max_y if max_y > canvas_height else 1)
        print("Factor de escala: ",scale_factor)
        scale_factor=scale_factor - 0.1
        # Escalar los puntos para que se ajusten al canvas
        if scale_factor < 1:
            points_g = points
            print("sin escalar: ", points_g)
            points = [(x * scale_factor, y * scale_factor) for x, y in points]
            print("escalado: ",points)

        # Dibujar la figura ajustada
        if tipo == "L":
            for i, point in enumerate(points):
                x1, y1 = point
                x2, y2 = points[(i + 1) % len(points)] if i + 1 < len(points) else (None, None)
                canvas.create_oval(x1 - 3, y1 - 3, x1 + 3, y1 + 3, fill="black")  # Dibuja puntos

                if x2 is not None and y2 is not None:
                    canvas.create_line(x1, y1, x2, y2, fill="blue", width=2)

                    # Calcular la distancia usando points_g (valores originales, no escalados)
                    x1_orig, y1_orig = points_g[i]
                    x2_orig, y2_orig = points_g[(i + 1) % len(points_g)] if i + 1 < len(points_g) else (None, None)
                    distancia = round(math.sqrt((x2_orig - x1_orig) ** 2 + (y2_orig - y1_orig) ** 2), 1)

                    mid_x, mid_y = (x1 + x2) / 2, (y1 + y2) / 2
                    canvas.create_text(mid_x, mid_y, text=f"{distancia} cm", font=("Arial", 8), fill="black")
        else:
            # Dibujar el polígono con los puntos calculados
            for i, point in enumerate(points):
                x1, y1 = point
                x2, y2 = points[(i + 1) % len(points)] if i + 1 < len(points) else points[0]
                canvas.create_oval(x1 - 3, y1 - 3, x1 + 3, y1 + 3, fill="black")  # Dibuja puntos

                if x2 is not None and y2 is not None:
                    canvas.create_line(x1, y1, x2, y2, fill="blue", width=2)

                    # Calcular la distancia usando points_g (valores originales, no escalados)
                    x1_orig, y1_orig = points_g[i]
                    x2_orig, y2_orig = points_g[(i + 1) % len(points_g)] if i + 1 < len(points_g) else points_g[0]
                    distancia = round(math.sqrt((x2_orig - x1_orig) ** 2 + (y2_orig - y1_orig) ** 2), 1)

                    mid_x, mid_y = (x1 + x2) / 2, (y1 + y2) / 2
                    canvas.create_text(mid_x, mid_y, text=f"{distancia} cm", font=("Arial", 8), fill="black")

    def editar_geometria():
        """
        Abre una ventana para editar las dimensiones de la figura seleccionada en el canvas.
        Actualiza los puntos y redibuja la figura en el canvas de la ventana principal.
        """
        figura = geometry_var.get()  # Obtiene el tipo de figura seleccionada
        nueva_ventana = tk.Toplevel(root)
        nueva_ventana.title(f"Editar {figura}")
        nueva_ventana.geometry("300x300")

        tk.Label(nueva_ventana, text=f"Editar dimensiones para {figura}", font=("Arial", 12, "bold")).pack(pady=10)

        campos = []

        def obtener_campos(figura):
            """Genera los campos según la figura seleccionada."""
            if figura == "L":
                return [("Segmento Vertical (cm)", 150), ("Segmento Horizontal (cm)", 50)]
            elif figura == "Rectángulo":
                return [("Base (cm)", 100), ("Altura (cm)", 150)]
            elif figura == "Triángulo":
                return [("Segmento 1 (cm)", 100), ("Segmento 2 (cm)", 100)]
            elif figura == "Línea":
                return [("Longitud (cm)", 150)]
            elif figura == "Circunferencia":
                return [("Diámetro (cm)", 100)]
            return []

        # Crear campos para edición basados en la figura seleccionada
        campos_figura = obtener_campos(figura)

        for i, (texto, valor) in enumerate(campos_figura):
            tk.Label(nueva_ventana, text=texto).pack(pady=5)
            entry = tk.Entry(nueva_ventana)
            entry.insert(0, valor)
            entry.pack(pady=5)
            campos.append(entry)

        def guardar_cambios():
            try:
                nuevos_valores = [float(campo.get()) for campo in campos]
                dibujar_figura(canvas, figura, nuevos_valores)
                nueva_ventana.destroy()
                messagebox.showinfo("Éxito", "Los valores se han actualizado correctamente.")
            except ValueError:
                messagebox.showerror("Error", "Ingrese valores numéricos válidos.")

        # Botón para guardar cambios
        tk.Button(nueva_ventana, text="Guardar", command=guardar_cambios).pack(pady=20)

        def cancelar():
            """Cierra la ventana sin realizar cambios."""
            nueva_ventana.destroy()

        # Botón para cancelar
        tk.Button(nueva_ventana, text="Cancelar", command=cancelar).pack(pady=5)

    def guardar_geometria():
        """
        Guarda los datos de la geometría seleccionada en la hoja "Área" del archivo Excel.
        Si la hoja ya contiene datos, sobrescribe la fila existente.
        """
        global points_g  # Usar la variable global points_g

        if not manager.archivo_cargado():
            messagebox.showwarning("Advertencia", "No hay ningún archivo abierto para guardar los datos.")
            return

        try:
            # Acceder o crear la hoja "Área"
            if "Área" not in manager.workbook.sheetnames:
                area_sheet = manager.workbook.create_sheet("Área")
                # Crear encabezados en la hoja
                encabezados = ["Tipo de Geometría", "Puntos", "Longitudes de Segmentos",
                               "Perímetro", "Área", "Radio Equivalente"]
                for col, encabezado in enumerate(encabezados, start=1):
                    area_sheet.cell(row=1, column=col, value=encabezado)
                fila_datos = 2  # Segunda fila para los datos
            else:
                area_sheet = manager.workbook["Área"]
                fila_datos = 2  # Sobrescribir siempre en la segunda fila

            # Usar la variable points_g para guardar los puntos originales
            geometry_type = geometry_var.get()
            if not points_g or len(points_g) < 2:
                messagebox.showerror("Error", "Debe haber al menos dos puntos para guardar la geometría.")
                return
            print("Points_g (guardar): ",points_g)

            # Calcular longitudes, perímetro y área
            longitudes, perimetro = calcular_longitudes_y_perimetro(points_g, geometry_type)
            area = calcular_area(geometry_type, longitudes)
            radio_equivalente = round(math.sqrt(area / math.pi), 4)


            # Guardar en la hoja Excel
            area_sheet.cell(row=fila_datos, column=1, value=geometry_type)
            area_sheet.cell(row=fila_datos, column=2, value=str(points_g))
            area_sheet.cell(row=fila_datos, column=3, value=str(longitudes))
            area_sheet.cell(row=fila_datos, column=4, value=round(perimetro, 2))
            area_sheet.cell(row=fila_datos, column=5, value=round(area, 4))  # Guardar área
            area_sheet.cell(row=fila_datos, column=6, value=radio_equivalente)  # Guardar radio equivalente

            # Guardar el archivo
            manager.guardar_archivo()
            messagebox.showinfo("Éxito", "Los datos de la geometría han sido guardados correctamente.")

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar la geometría: {e}")

    def calcular_area(tipo_geometria, longitudes):
        """
        Calcula el área de la figura según su tipo.

        :param tipo_geometria: Tipo de geometría ('Rectángulo', 'Triángulo', 'Línea', 'L').
        :param longitudes: Longitudes de los segmentos.
        :return: Área calculada de la figura.
        """
        if tipo_geometria == "Rectángulo":
            # Área = Base × Altura
            base, altura = longitudes[:2]
            return (base * altura)/10000


        elif tipo_geometria == "Línea":
            # El área de una línea es igual a su longitud total
            return sum(longitudes)/100

        elif tipo_geometria == "L":
            # Área = suma de los segmentos (vertical + horizontal)
            return sum(longitudes[:2])/100

        else:
            return 0
    figura_desde_excel = False  # Indicar que ya no proviene de Excel


    # Vincular evento de cambio en combobox para dibujar automáticamente
    geometry_combo.bind("<<ComboboxSelected>>", lambda e: on_geometry_change)

    # Dibujar la geometría inicial
    draw_geometry()


# Función para abrir la ventana de "Datos de proyecto"
def abrir_datos_proyecto():
    """
    Abre una ventana para mostrar y modificar los datos del proyecto.
    Utiliza una instancia de ExcelManager para interactuar con el archivo Excel.
    """
    global file_path

    # Función para cargar datos existentes en los Entry
    def cargar_datos_existentes():
        if "Información" in manager.workbook.sheetnames:
            info_sheet = manager.workbook["Información"]
            if info_sheet.cell(row=2, column=1).value:  # Verificar si hay datos en la fila 2
                entry_titulo.insert(0, info_sheet.cell(row=2, column=1).value)
                combo_provincia.set(info_sheet.cell(row=2, column=2).value)
                entry_canton.insert(0, info_sheet.cell(row=2, column=3).value)
                entry_calle_principal.insert(0, info_sheet.cell(row=2, column=4).value)
                entry_calle_secundaria.insert(0, info_sheet.cell(row=2, column=5).value)
                entry_referencia.insert(0, info_sheet.cell(row=2, column=6).value)
                fecha_inicio.set_date(info_sheet.cell(row=2, column=7).value)
                fecha_final.set_date(info_sheet.cell(row=2, column=8).value)
                entry_nombre_dueno.insert(0, info_sheet.cell(row=2, column=9).value)
                entry_id_dueno.insert(0, info_sheet.cell(row=2, column=10).value)
                entry_nombre_profesional.insert(0, info_sheet.cell(row=2, column=11).value)
                entry_id_profesional.insert(0, info_sheet.cell(row=2, column=12).value)
                reg_senescyt.insert(0, info_sheet.cell(row=2, column=13).value)
                entry_longitud.insert(0, info_sheet.cell(row=2, column=14).value)
                entry_latitud.insert(0, info_sheet.cell(row=2, column=15).value)

                # Verificar si hay imagen
                if info_sheet.cell(row=2, column=16).value == "Imagen insertada":
                    ubicacion_mapa.config(text="Imagen cargada previamente", fg="green")
                else:
                    ubicacion_mapa.config(text="No hay ningún archivo cargado.", fg="red")

            else:
                ubicacion_mapa.config(text="No hay ningún archivo cargado.", fg="red")
    # Función para cargar imagen satelital
    def cargar_imagen():
        global file_path
        file_path = filedialog.askopenfilename(title="Cargar Imagen Satelital", filetypes=[("Images", "*.png *.jpg")])

        if file_path:
            ubicacion_mapa.config(text="Cargado con éxito", fg="green")
        else:
            file_path = None

    if not manager.archivo_cargado():
        messagebox.showwarning("Advertencia", "No hay ningún archivo abierto o cargado.")
        return

    # Crear la ventana de datos de proyecto
    ventana_datos = tk.Toplevel(root)
    ventana_datos.title("Datos de proyecto")
    ventana_datos.geometry("570x395")


    # Función para guardar los datos
    def guardar_datos():
        global file_patch

        try:
            if "Información" not in manager.workbook.sheetnames:
                info_sheet = manager.workbook.create_sheet("Información")
                encabezados = [
                    "Título del Proyecto", "Provincia", "Cantón", "Calle Principal",
                    "Calle Secundaria", "Referencia", "Fecha de Inicio", "Fecha de Finalización",
                    "Nombre Dueño", "ID Dueño", "Nombre Profesional", "ID Profesional",
                    "Reg. SENESCYT", "Longitud", "Latitud", "Ruta Imagen Satelital"
                ]
                col=16
                for col, encabezado in enumerate(encabezados, start=1):
                    info_sheet.cell(row=1, column=col, value=encabezado)
            else:
                info_sheet = manager.workbook["Información"]

            fila = 2  # Sobrescribir datos en fila 2
            datos = [
                entry_titulo.get(), combo_provincia.get(), entry_canton.get(), entry_calle_principal.get(),
                entry_calle_secundaria.get(), entry_referencia.get(), fecha_inicio.get_date(), fecha_final.get_date(),
                entry_nombre_dueno.get(), entry_id_dueno.get(), entry_nombre_profesional.get(),
                entry_id_profesional.get(),
                reg_senescyt.get(), entry_longitud.get(), entry_latitud.get(), "Imagen insertada"
                # Añadir la ruta de la imagen
            ]

            for col, valor in enumerate(datos, start=1):
                info_sheet.cell(row=fila, column=col, value=valor)
            # Eliminar imágenes existentes

            # Insertar la imagen en el Excel
            if file_path:
                info_sheet._images.clear()
                img = Image(file_path)
                img.width, img.height = 300, 300  # Redimensionar la imagen
                info_sheet.add_image(img, "P2")  # Insertar la imagen en la celda P2

            # Guardar el archivo Excel
            manager.guardar_archivo()
            messagebox.showinfo("Guardado", "Los datos se han guardado correctamente.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar: {e}")


    # Función para limpiar los datos ingresados
    def limpiar_datos():
        for entry in entradas.values():
            entry.delete(0, tk.END)
        combo_provincia.set("")
        fecha_inicio.set_date("")
        fecha_final.set_date("")
        estado_imagen.config(text="No cargado", fg="red")
        estado_circulo.itemconfig(indicador, fill="red")


    # --- Entradas y Widgets ---
    entradas = {}
    # Frame - Detalles del Proyecto
    frame_detalles = tk.LabelFrame(ventana_datos, text="Detalles del Proyecto", padx=5, pady=5)
    frame_detalles.grid(column=0, row=0, columnspan=2, pady=5, padx=10, rowspan=3)

    tk.Label(frame_detalles, text="Título del Proyecto:").grid(row=0, column=0, sticky="e", pady=5)
    entry_titulo = tk.Entry(frame_detalles)
    entry_titulo.grid(row=0, column=1, pady=5)

    tk.Label(frame_detalles, text="Provincia:").grid(row=1, column=0, sticky="e")
    combo_provincia = ttk.Combobox(frame_detalles, values=[
        "Azuay", "Bolívar", "Cañar", "Carchi", "Chimborazo", "Cotopaxi", "El Oro", "Esmeraldas",
        "Galápagos", "Guayas", "Imbabura", "Loja", "Los Ríos", "Manabí", "Morona Santiago", "Napo", "Orellana",
        "Pastaza", "Pichincha", "Santa Elena", "Santo Domingo de los Tsáchilas", "Sucumbíos", "Tungurahua",
        "Zamora Chinchipe"
    ])
    combo_provincia.grid(row=1, column=1, padx=5, pady=5)

    tk.Label(frame_detalles, text="Cantón:").grid(row=2, column=0, sticky="e")
    entry_canton = tk.Entry(frame_detalles)
    entry_canton.grid(row=2, column=1, pady=5)

    tk.Label(frame_detalles, text="Calle Principal:").grid(row=3, column=0, sticky="e")
    entry_calle_principal = tk.Entry(frame_detalles)
    entry_calle_principal.grid(row=3, column=1, pady=5)

    tk.Label(frame_detalles, text="Calle Secundaria:").grid(row=4, column=0, sticky="e")
    entry_calle_secundaria = tk.Entry(frame_detalles)
    entry_calle_secundaria.grid(row=4, column=1, pady=5)

    tk.Label(frame_detalles, text="Referencia:").grid(row=5, column=0, sticky="e")
    entry_referencia = tk.Entry(frame_detalles)
    entry_referencia.grid(row=5, column=1, pady=5)

    tk.Label(frame_detalles, text="Fecha de inicio").grid(row=6, column=0, sticky="e")
    fecha_inicio = DateEntry(frame_detalles)
    fecha_inicio.grid(row=6, column=1, pady=5)

    tk.Label(frame_detalles, text="Fecha de finalización").grid(row=7, column=0, sticky="e")
    fecha_final = DateEntry(frame_detalles)
    fecha_final.grid(row=7, column=1, pady=5)

    # Frame - Coordenadas
    frame_utm = tk.LabelFrame(ventana_datos, text="Coordenadas UTM")
    frame_utm.grid(column=0, row=3, columnspan=2, pady=5, padx=10,rowspan=2, sticky="ew")
    entry_longitud = tk.Entry(frame_utm)
    entry_latitud = tk.Entry(frame_utm)
    tk.Label(frame_utm, text="Longitud:").grid(row=0, column=0,pady=5, padx=5, sticky="w")
    entry_longitud.grid(row=0, column=1,pady=5, padx=5, sticky="ew")
    tk.Label(frame_utm, text="Latitud:").grid(row=1, column=0,pady=5, padx=5, sticky="w")
    entry_latitud.grid(row=1, column=1,pady=5, padx=5, sticky="ew")

    # Frame - Dueño del Proyecto
    frame_dueno = tk.LabelFrame(ventana_datos, text="Dueño del Proyecto")
    frame_dueno.grid(column=2, row=0, columnspan=2, pady=5, padx=10)

    entry_nombre_dueno = tk.Entry(frame_dueno)
    entry_nombre_dueno.grid(row=0, column=1,pady=5, padx=5)
    tk.Label(frame_dueno, text="Nombre:").grid(row=0, column=0, sticky="e" ,pady=5, padx=5)

    entry_id_dueno = tk.Entry(frame_dueno)
    entry_id_dueno.grid(row=1, column=1,pady=5, padx=5)
    tk.Label(frame_dueno, text="Documento ID:").grid(row=1, column=0, sticky="e" ,pady=5, padx=5)

    # Frame - Profesional Responsable
    frame_profesional = tk.LabelFrame(ventana_datos, text="Profesional Responsable")
    frame_profesional.grid(column=2, row=1, columnspan=2, pady=5, padx=10)

    entry_nombre_profesional = tk.Entry(frame_profesional)
    entry_nombre_profesional.grid(row=0, column=1, pady=5, padx=5)
    tk.Label(frame_profesional, text="Nombre:").grid(row=0, column=0, sticky="e", pady=5, padx=5)

    entry_id_profesional = tk.Entry(frame_profesional)
    entry_id_profesional.grid(row=1, column=1, pady=5, padx=5)
    tk.Label(frame_profesional, text="Documento ID:").grid(row=1, column=0, sticky="e", pady=5, padx=5)

    reg_senescyt = tk.Entry(frame_profesional)
    reg_senescyt.grid(row=2, column=1, pady=5, padx=5)
    tk.Label(frame_profesional, text="Reg. SENESCYT:").grid(row=2, column=0, sticky="e", pady=5, padx=5)


    # Frame - Imagen
    frame_imagen = tk.LabelFrame(ventana_datos, text="Cargar imagen satelital")
    frame_imagen.grid(column=2, row=2, columnspan=2, pady=5, padx=10, sticky="ew")
    tk.Button(frame_imagen, text="Examinar", command=cargar_imagen).grid(row=0, column=0, pady=5, padx=5)

    ubicacion_mapa = tk.Label(frame_imagen, pady=5, padx=5)
    ubicacion_mapa.grid(row=0, column=1, pady=10, padx=5)
    ubicacion_mapa.configure(fg="green")

    # Frame - Botones
    frame_botones = tk.Label(ventana_datos)
    frame_botones.grid(row=3, column=2, columnspan=3, pady=5, padx=5, sticky="ew", rowspan=2)

    cargar_datos_existentes()

    tk.Button(frame_botones, text="Guardar", command=guardar_datos, pady=5, padx=10).grid(row=0, column=0)
    tk.Button(frame_botones, text="Limpiar", command=limpiar_datos, pady=5, padx=10).grid(row=0, column=1)
    tk.Button(frame_botones, text="Cancelar", command=ventana_datos.destroy, pady=5, padx=10).grid(row=0, column=2)

# Función para abrir la ventana "Datos de Resistencia"
def abrir_datos_resistencia():
    ventana = tk.Toplevel(root)
    ventana.title("Línea con Perfiles y Datos de Resistencia")
    ventana.geometry("700x450")  # Ajusta el tamaño según lo necesites

    # Dimensiones del Canvas
    ancho_canvas = 150
    alto_canvas = 150

    # Crear el Canvas
    canvas = tk.Canvas(ventana, width=ancho_canvas, height=alto_canvas, bg="white")
    canvas.grid(row=0, column=0, columnspan=5, pady=10)

    # Variables para almacenar el perfil seleccionado y entradas
    perfil_seleccionado = tk.StringVar()
    entradas = []

    # Función para dibujar la línea según el perfil seleccionado
    def dibujar_linea_perfil(perfil):
        perfil_seleccionado.set(perfil)  # Guardar el perfil seleccionado
        canvas.delete("all")

        y_mitad = alto_canvas // 2
        x_mitad = ancho_canvas // 2
        x_inicio = 0 + 7
        x_final = ancho_canvas - 4
        radio_punto = 4

        if perfil == "Ruta 1":
            canvas.create_line(x_inicio, y_mitad, x_final, y_mitad, fill="black", width=2)
            canvas.create_oval(x_inicio - radio_punto, y_mitad - radio_punto, x_inicio + radio_punto,
                               y_mitad + radio_punto, fill="red")
            canvas.create_oval(x_final - radio_punto, y_mitad - radio_punto, x_final + radio_punto,
                               y_mitad + radio_punto, fill="red")
        elif perfil == "Ruta 2":
            canvas.create_line(x_mitad, 0 + 7, x_mitad, alto_canvas - 4, fill="black", width=2)
            canvas.create_oval(x_mitad - radio_punto, 7 - radio_punto, x_mitad + radio_punto, 7 + radio_punto,
                               fill="red")
            canvas.create_oval(x_mitad - radio_punto, alto_canvas - 4 - radio_punto, x_mitad + radio_punto,
                               alto_canvas - 4 + radio_punto, fill="red")
        elif perfil == "Ruta 3":
            canvas.create_line(7, 7, ancho_canvas - 4, alto_canvas - 4, fill="black", width=2)
            canvas.create_oval(7 - radio_punto, 7 - radio_punto, 7 + radio_punto, 7 + radio_punto, fill="red")
            canvas.create_oval(ancho_canvas - 4 - radio_punto, alto_canvas - 4 - radio_punto,
                               ancho_canvas - 4 + radio_punto, alto_canvas - 4 + radio_punto, fill="red")
        elif perfil == "Ruta 4":
            canvas.create_line(7, alto_canvas - 4, ancho_canvas - 4, 7, fill="black", width=2)
            canvas.create_oval(7 - radio_punto, alto_canvas - 4 - radio_punto, 7 + radio_punto,
                               alto_canvas - 4 + radio_punto, fill="red")
            canvas.create_oval(ancho_canvas - 4 - radio_punto, 7 - radio_punto, ancho_canvas - 4 + radio_punto,
                               7 + radio_punto, fill="red")

    # Crear el menú contextual para perfiles
    menu_contextual_perfiles = tk.Menu(ventana, tearoff=0)
    menu_contextual_perfiles.add_command(label="Ruta 1", command=lambda: dibujar_linea_perfil("Ruta 1"))
    menu_contextual_perfiles.add_command(label="Ruta 2", command=lambda: dibujar_linea_perfil("Ruta 2"))
    menu_contextual_perfiles.add_command(label="Ruta 3", command=lambda: dibujar_linea_perfil("Ruta 3"))
    menu_contextual_perfiles.add_command(label="Ruta 4", command=lambda: dibujar_linea_perfil("Ruta 4"))

    # Botón para abrir el menú contextual
    btn_dibujar_perfil = tk.Button(ventana, text="Dibujar Perfil")
    btn_dibujar_perfil.grid(row=1, column=0, columnspan=5, pady=10)
    btn_dibujar_perfil.bind("<Button-1>", lambda event: menu_contextual_perfiles.post(event.x_root, event.y_root))

    # Etiquetas de encabezado de columnas
    columnas = ["1m", "2m", "3m", "4m"]
    for i, columna in enumerate(columnas):
        tk.Label(ventana, text=columna).grid(row=2, column=i + 1, padx=5, pady=5)

    # Etiqueta de fila "Resistencia" y campos de entrada
    tk.Label(ventana, text="Resistencia").grid(row=3, column=0, padx=10, pady=5)
    for columna in range(1, 5):
        entrada = tk.Entry(ventana, width=5)
        entrada.grid(row=3, column=columna, padx=5, pady=5)
        entradas.append(entrada)

    # Crear la tabla para mostrar los datos guardados
    tabla = ttk.Treeview(ventana, columns=("Perfil", "1m", "2m", "3m", "4m"), show="headings", height=4)
    tabla.grid(row=0, column=6, rowspan=1, padx=5, pady=5, sticky='e')

    # Configurar encabezados de la tabla
    tabla.heading("Perfil", text="Perfil")
    tabla.heading("1m", text="1m")
    tabla.heading("2m", text="2m")
    tabla.heading("3m", text="3m")
    tabla.heading("4m", text="4m")

    # Ajustar ancho de cada columna de la tabla
    tabla.column("Perfil", width=50, anchor="center")
    tabla.column("1m", width=30, anchor="center")
    tabla.column("2m", width=30, anchor="center")
    tabla.column("3m", width=30, anchor="center")
    tabla.column("4m", width=30, anchor="center")

    # Función para cargar los datos existentes en el archivo actual a la tabla
    def cargar_datos_existentes(tabla):
        """Carga los datos existentes desde la hoja 'Resistencias' del archivo Excel a la tabla."""
        if not manager.archivo_cargado():
            messagebox.showwarning("Advertencia", "No hay ningún archivo abierto o cargado.")
            return

        try:
            # Verificar si la hoja "Resistencias" existe
            if "Resistencias" not in manager.workbook.sheetnames:
                messagebox.showinfo("Información", "No se encontró la hoja 'Resistencias' en el archivo.")
                return

            # Acceder a la hoja "Resistencias"
            resistencias_sheet = manager.workbook["Resistencias"]

            # Obtener encabezados de la primera fila
            encabezados = [cell.value for cell in resistencias_sheet[1] if cell.value]
            if not encabezados or encabezados[0] != "Perfil":
                messagebox.showerror("Error", "Los encabezados de la hoja 'Resistencias' no son válidos.")
                return

            # Limpiar la tabla antes de cargar nuevos datos
            for item in tabla.get_children():
                tabla.delete(item)

            # Configurar columnas dinámicamente en la tabla
            tabla["columns"] = encabezados
            for col in encabezados:
                tabla.heading(col, text=col)
                tabla.column(col, width=80, anchor="center")

            # Leer los datos de la hoja (a partir de la segunda fila)
            for row in resistencias_sheet.iter_rows(min_row=2, values_only=True):
                # Insertar solo si hay datos válidos
                if any(cell not in [None, ""] for cell in row):
                    tabla.insert("", "end", values=row)

        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar los datos de la hoja 'Resistencias': {e}")

    # Llamar a la función para cargar datos existentes en la tabla al abrir la ventana
    cargar_datos_existentes(tabla)

    # Función para actualizar el canvas cuando se selecciona una fila
    def actualizar_canvas(event):
        selected_item = tabla.selection()
        if selected_item:
            valores = tabla.item(selected_item)["values"]
            if valores:
                perfil = valores[0]  # El primer valor es el perfil (Ruta 1, Ruta 2, etc.)
                dibujar_linea_perfil(perfil)

    # Vincular el evento de selección en la tabla
    tabla.bind("<<TreeviewSelect>>", actualizar_canvas)

    # Función para eliminar una fila seleccionada de la tabla y el archivo Excel
    def eliminar_fila():
        if not manager.archivo_cargado():
            messagebox.showwarning("Advertencia", "No hay ningún archivo abierto o cargado.")
            return

        selected_item = tabla.selection()
        if not selected_item:
            messagebox.showwarning("Advertencia", "Seleccione una fila para eliminar.")
            return
        try:
            # Obtener los datos de la fila seleccionada
            perfil_data = tabla.item(selected_item)["values"]
            perfil = perfil_data[0]  # El perfil (e.g., Horizontal, Vertical, etc.)

            # Mapa de perfiles y filas asociadas en el Excel
            perfiles_filas = {
                "Ruta 1": 21,
                "Ruta 2": 22,
                "Ruta 3": 23,
                "Ruta 4": 24
            }

            if perfil in perfiles_filas:
                fila = perfiles_filas[perfil]

                # Eliminar datos de la fila en el archivo Excel
                manager.escribir_dato(f"A{fila}", None)  # Limpia la celda del perfil
                for col in range(2, 6):  # Limpia las columnas B a E
                    manager.escribir_dato(f"{chr(64 + col)}{fila}", None)

                # Guardar cambios en el archivo
                manager.guardar_archivo()

                # Eliminar la fila de la tabla en la interfaz
                tabla.delete(selected_item)
                messagebox.showinfo("Eliminado", "La fila ha sido eliminada del archivo y la tabla.")
            else:
                messagebox.showwarning("Advertencia", "El perfil seleccionado no es válido.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo eliminar la fila: {e}")

    # Botón para eliminar la fila seleccionada
    btn_eliminar = tk.Button(ventana, text="Eliminar Fila Seleccionada", command=eliminar_fila)
    btn_eliminar.grid(row=6, column=0, columnspan=5, pady=10)

    # Función para validar y añadir datos en la tabla
    def anadir_datos_perfil():
        perfil = perfil_seleccionado.get()
        if not perfil:
            messagebox.showerror("Error", "Seleccione un perfil antes de ingresar valores de resistencia.")
            return

        # Recoger los valores de resistencia y validar que sean numéricos o vacíos
        resistencias = []
        for entrada in entradas:
            valor = entrada.get()
            try:
                resistencias.append(float(valor) if valor else "")  # Convertir a float si no está vacío
            except ValueError:
                messagebox.showerror("Error", "Ingrese valores válidos en los campos de resistencia.")
                return

        # Buscar si el perfil y la distancia ya existen en la tabla
        for item in tabla.get_children():
            item_data = tabla.item(item)["values"]
            if item_data[0] == perfil:  # Si el perfil ya está en la tabla
                # Actualizar los valores de la tabla, incluyendo vacíos
                for i in range(1, 5):  # Columnas de distancia (1m a 4m)
                    item_data[i] = resistencias[i - 1]  # Actualizar con el valor ingresado o vacío
                tabla.item(item, values=item_data)  # Reemplazar valores existentes en la tabla
                return

        # Si el perfil no existe en la tabla, añadirlo
        tabla.insert("", "end", values=(perfil, *resistencias))

    # Botón para añadir los datos a la tabla
    btn_anadir_datos = tk.Button(ventana, text="Añadir Datos Perfil", command=anadir_datos_perfil)
    btn_anadir_datos.grid(row=4, column=0, columnspan=5, pady=10)

    # Función para guardar los datos de la tabla en el archivo Excel
    def guardar_datos_resistencia():
        if not manager.archivo_cargado():
            messagebox.showwarning("Advertencia", "No hay ningún archivo abierto para guardar los datos.")
            return

        try:
            # Verificar si la hoja "Resistencias" existe, y si no, crearla
            if "Resistencias" not in manager.workbook.sheetnames:
                resistencias_sheet = manager.workbook.create_sheet("Resistencias")
            else:
                resistencias_sheet = manager.workbook["Resistencias"]

            # Limpiar la hoja actual antes de escribir los nuevos datos
            resistencias_sheet.delete_rows(1, resistencias_sheet.max_row)

            # Obtener los datos de la tabla y filtrar columnas vacías o None
            datos_tabla = []
            for item in tabla.get_children():
                fila = tabla.item(item)["values"]
                datos_tabla.append(fila)

            # Mapeo de columnas originales a distancias
            distancias_originales = ["Perfil", "1m", "2m", "3m", "4m"]

            # Determinar qué columnas contienen datos válidos (excluir None y '')
            columnas_validas = [0]  # La columna "Perfil" siempre se guarda
            for col in range(1, len(datos_tabla[0])):  # Excluir la columna "Perfil"
                if any(fila[col] not in [None, ""] for fila in datos_tabla):
                    columnas_validas.append(col)

            # Crear la lista de encabezados filtrados con las distancias correctas
            encabezados_filtrados = [distancias_originales[i] for i in columnas_validas]

            # Escribir los encabezados en la hoja Excel
            for col, encabezado in enumerate(encabezados_filtrados, start=1):
                resistencias_sheet.cell(row=1, column=col, value=encabezado)

            # Escribir los datos en la hoja filtrando columnas vacías y None
            for i, fila in enumerate(datos_tabla, start=2):
                for col_idx, col in enumerate(columnas_validas):
                    valor = fila[col]
                    resistencias_sheet.cell(row=i, column=col_idx + 1, value=valor if valor not in [None, ""] else "")

            # Guardar los cambios en el archivo Excel
            manager.guardar_archivo()
            messagebox.showinfo("Guardado", "Los datos de resistencia se han guardado correctamente.")

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar los datos de resistencia: {e}")

        # Actualizar la ventana principal
        actualizar_datos_principal()

    # Botón para guardar los datos en el Excel
    btn_guardar = tk.Button(ventana, text="Guardar", command=guardar_datos_resistencia)
    btn_guardar.grid(row=5, column=0, columnspan=5, pady=10)


def abrir_datos_transformador():
    if not manager.archivo_cargado():
        messagebox.showwarning("Advertencia", "No hay ningún archivo abierto o cargado.")
        return

        # Crear la ventana "Datos del transformador"
    ventana_transformador = tk.Toplevel(root)
    ventana_transformador.title("Datos del transformador")
    ventana_transformador.geometry("400x600")

    # Campos de entrada con etiquetas para cada dato solicitado
    tk.Label(ventana_transformador, text="Tensión del primario (V)").pack(pady=5)
    tension_primario_entry = tk.Entry(ventana_transformador)
    tension_primario_entry.pack(pady=5)

    tk.Label(ventana_transformador, text="Tensión del secundario (V)").pack(pady=5)
    tension_secundario_entry = tk.Entry(ventana_transformador)
    tension_secundario_entry.pack(pady=5)

    tk.Label(ventana_transformador, text="Potencia nominal del transformador (kVA)").pack(pady=5)
    potencia_nominal_entry = tk.Entry(ventana_transformador)
    potencia_nominal_entry.pack(pady=5)

    tk.Label(ventana_transformador, text="Configuración del devanado").pack(pady=5)
    configuracion_devanado_var = tk.StringVar(value="Y-Estrella")
    configuracion_devanado_menu = ttk.Combobox(
        ventana_transformador,
        textvariable=configuracion_devanado_var,
        values=["Y-Estrella", "Δ-Delta"],
        state="readonly"
    )
    configuracion_devanado_menu.pack(pady=5)

    tk.Label(ventana_transformador, text="Tipo de conexión a tierra").pack(pady=5)
    tipo_conexion_tierra_var = tk.StringVar(value="Neutro directo a tierra")
    tipo_conexion_tierra_menu = ttk.Combobox(
        ventana_transformador,
        textvariable=tipo_conexion_tierra_var,
        values=["Neutro directo a tierra", "Neutro con resistencia", "Neutro aislado"],
        state="readonly"
    )
    tipo_conexion_tierra_menu.pack(pady=5)

    tk.Label(ventana_transformador, text="Impedancia de cortocircuito (%) o (pu)").pack(pady=5)
    frame_impedancia = tk.Frame(ventana_transformador)
    frame_impedancia.pack(pady=5)
    impedancia_entry = tk.Entry(frame_impedancia, width=15)
    impedancia_entry.pack(side="left")
    impedancia_var = tk.StringVar(value="%")
    impedancia_menu = ttk.Combobox(
        frame_impedancia,
        textvariable=impedancia_var,
        values=["%", "pu"],
        state="readonly",
        width=5
    )
    impedancia_menu.pack(side="left", padx=5)

    tk.Label(ventana_transformador, text="Corriente de falla a tierra (A)").pack(pady=5)
    corriente_falla_entry = tk.Entry(ventana_transformador)
    corriente_falla_entry.pack(pady=5)

    tk.Label(ventana_transformador, text="Voltaje Línea a Línea (V)").pack(pady=5)
    voltaje_linea_linea_entry = tk.Entry(ventana_transformador)
    voltaje_linea_linea_entry.pack(pady=5)

    tk.Label(ventana_transformador, text="Voltaje Línea a Neutro (V)").pack(pady=5)
    voltaje_linea_neutro_entry = tk.Entry(ventana_transformador)
    voltaje_linea_neutro_entry.pack(pady=5)

    # Función para cargar datos desde el archivo actual en los campos de la ventana
    def cargar_datos_transformador():
        try:
            # Verificar si la hoja "trafos" existe
            if "trafos" in manager.workbook.sheetnames:
                trafos_sheet = manager.workbook["trafos"]

                # Verificar si hay datos en la hoja (después de los encabezados)
                if trafos_sheet.max_row > 1:
                    ultima_fila = trafos_sheet.max_row
                    tension_primario_entry.insert(0, trafos_sheet.cell(row=ultima_fila, column=1).value or "")
                    tension_secundario_entry.insert(0, trafos_sheet.cell(row=ultima_fila, column=2).value or "")
                    potencia_nominal_entry.insert(0, trafos_sheet.cell(row=ultima_fila, column=3).value or "")
                    configuracion_devanado_var.set(trafos_sheet.cell(row=ultima_fila, column=4).value or "Y-Estrella")
                    tipo_conexion_tierra_var.set(
                        trafos_sheet.cell(row=ultima_fila, column=5).value or "Neutro directo a tierra")
                    impedancia = trafos_sheet.cell(row=ultima_fila, column=6).value or ""
                    if impedancia:
                        # Separar valor y unidad de la impedancia
                        if " " in impedancia:
                            valor, unidad = impedancia.split(" ")
                            impedancia_entry.insert(0, valor)
                            impedancia_var.set(unidad)
                        else:
                            impedancia_entry.insert(0, impedancia)
                    corriente_falla_entry.insert(0, trafos_sheet.cell(row=ultima_fila, column=7).value or "")
                else:
                    messagebox.showinfo("Información", "No hay datos previos de transformadores para cargar.")
            else:
                messagebox.showinfo("Información", "No se encontró la hoja 'trafos' en el archivo Excel.")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo cargar los datos: {e}")


    # Función para guardar los datos ingresados en el archivo Excel
    def guardar_datos_transformador():
        if not manager.archivo_cargado():
            messagebox.showwarning("Advertencia", "No hay ningún archivo abierto para guardar los datos.")
            return

        try:
            # Acceder o crear la hoja "trafos"
            if "trafos" not in manager.workbook.sheetnames:
                trafos_sheet = manager.workbook.create_sheet("trafos")
                trafos_sheet.append([
                    "Tensión Primario (V)",
                    "Tensión Secundario (V)",
                    "Potencia Nominal (kVA)",
                    "Configuración Devanado",
                    "Tipo Conexión Tierra",
                    "Impedancia",
                    "Corriente Falla (A)"
                ])
            else:
                trafos_sheet = manager.workbook["trafos"]

            # Guardar datos en la hoja "trafos"
            datos = [
                tension_primario_entry.get(),
                tension_secundario_entry.get(),
                potencia_nominal_entry.get(),
                configuracion_devanado_var.get(),
                tipo_conexion_tierra_var.get(),
                f"{impedancia_entry.get()} {impedancia_var.get()}",
                corriente_falla_entry.get()
            ]
            trafos_sheet.append(datos)
            manager.guardar_archivo()

            messagebox.showinfo("Guardado", "Los datos del transformador se han guardado en la hoja 'trafos'.")
            ventana_transformador.destroy()

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar los datos: {e}")

    # Función para limpiar los campos de entrada
    def limpiar_campos():
        try:
            # Limpiar todos los campos de entrada
            tension_primario_entry.delete(0, tk.END)
            tension_secundario_entry.delete(0, tk.END)
            potencia_nominal_entry.delete(0, tk.END)
            configuracion_devanado_var.set("Y-Estrella")
            tipo_conexion_tierra_var.set("Neutro directo a tierra")
            impedancia_entry.delete(0, tk.END)
            impedancia_var.set("%")
            corriente_falla_entry.delete(0, tk.END)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo limpiar los campos: {e}")

    # Botones de Guardar y Limpiar
    tk.Button(ventana_transformador, text="Guardar", command=guardar_datos_transformador).pack(pady=10)
    tk.Button(ventana_transformador, text="Limpiar", command=limpiar_campos).pack(pady=5)

    # Cargar datos del archivo al abrir la ventana
    cargar_datos_transformador()


def abrir_datos_conductor():
    # Crear la ventana para los datos de conductor
    ventana_conductor = tk.Toplevel(root)
    ventana_conductor.title("Datos de Conductor")
    ventana_conductor.geometry("430x470")
    ventana_conductor.resizable(False, False)

    # Frame para Datos de Varilla
    frame_varilla = tk.LabelFrame(ventana_conductor, text="Datos de Varilla", padx=10, pady=10)
    frame_varilla.pack(fill="x", padx=10, pady=10)


    # Datos ordenados descendentemente según el tamaño
    diametros = ["19mm", "18mm", "16mm", "14.7mm", "14mm", "12.5mm", "12mm", "3/4 in", "5/8 in"]

    tk.Label(frame_varilla, text="Diámetro").grid(row=0, column=0, sticky="w", pady=5)
    diametro_combo = ttk.Combobox(frame_varilla, values=diametros, state="readonly")
    diametro_combo.grid(row=0, column=1, pady=5)
    diametro_combo.set("5/8 in")

    tk.Label(frame_varilla, text="Longitud (cm)").grid(row=1, column=0, sticky="w", pady=5)
    longitud_combo = ttk.Combobox(frame_varilla, values=["120", "150", "180","240"], state="readonly")
    longitud_combo.set("240")  # Valor predeterminado
    longitud_combo.grid(row=1, column=1, pady=5)



    # Frame para Datos de Conductor
    frame_conductor = tk.LabelFrame(ventana_conductor, text="Datos de Conductor", padx=10, pady=10)
    frame_conductor.pack(fill="x", padx=10, pady=10)

    Materiales_tipo= [
        {
            "description": "Cobre, recocido, trefilado blando",
            "conductivity": 100,
            "alpha_20C": 0.00393,
            "K0_0C": 234,
            "fusion_temp": 1083,
            "resistivity_20C": 1.72,
            "thermal_capacity": 3.42,
            "Kf": 7
        },
        {
            "description": "Cobre, comercial, trefilado duro",
            "conductivity": 97,
            "alpha_20C": 0.00381,
            "K0_0C": 242,
            "fusion_temp": 1084,
            "resistivity_20C": 1.78,
            "thermal_capacity": 3.42,
            "Kf": 7.06
        },
        {
            "description": "Alambre de acero recubierto de cobre",
            "conductivity": 40,
            "alpha_20C": 0.00378,
            "K0_0C": 245,
            "fusion_temp": 1084,
            "resistivity_20C": 4.4,
            "thermal_capacity": 3.85,
            "Kf": 10.45
        },
        {
            "description": "Alambre de acero recubierto de cobre",
            "conductivity": 30,
            "alpha_20C": 0.00378,
            "K0_0C": 245,
            "fusion_temp": 1084,
            "resistivity_20C": 5.86,
            "thermal_capacity": 3.85,
            "Kf": 12.06
        },
        {
            "description": "Varilla de acero recubierta de cobre",
            "conductivity": 20,
            "alpha_20C": 0.00378,
            "K0_0C": 245,
            "fusion_temp": 1084,
            "resistivity_20C": 8.62,
            "thermal_capacity": 3.85,
            "Kf": 14.64
        },
        {
            "description": "Aluminio, grado EC",
            "conductivity": 61,
            "alpha_20C": 0.00403,
            "K0_0C": 228,
            "fusion_temp": 657,
            "resistivity_20C": 2.86,
            "thermal_capacity": 2.56,
            "Kf": 12.12
        },
        {
            "description": "Aluminio, aleación 5005",
            "conductivity": 53.5,
            "alpha_20C": 0.00353,
            "K0_0C": 263,
            "fusion_temp": 652,
            "resistivity_20C": 3.22,
            "thermal_capacity": 2.6,
            "Kf": 12.41
        },
        {
            "description": "Aluminio, aleación 6201",
            "conductivity": 52.5,
            "alpha_20C": 0.00347,
            "K0_0C": 268,
            "fusion_temp": 654,
            "resistivity_20C": 3.28,
            "thermal_capacity": 2.6,
            "Kf": 12.47
        },
        {
            "description": "Alambre de acero recubierto de aluminio",
            "conductivity": 20.3,
            "alpha_20C": 0.0036,
            "K0_0C": 258,
            "fusion_temp": 657,
            "resistivity_20C": 8.48,
            "thermal_capacity": 3.58,
            "Kf": 17.2
        }
    ]

    # Extraer las descripciones de los materiales
    descripciones = [material["description"] for material in Materiales_tipo]

    tk.Label(frame_conductor, text="Material").grid(row=0, column=0, sticky="w", pady=5)
    material_conductor_combo = ttk.Combobox(frame_conductor, values=descripciones, state="readonly")
    material_conductor_combo.grid(row=0, column=1, pady=5)
    material_conductor_combo.set(descripciones[0])

    conductor_data = [
        {
            "MCM": 350,
            "AWG": None,
            "nominal_area_mm2": 177.35,
            "diameter_m": 0.015
        },
        {
            "MCM": 300,
            "AWG": None,
            "nominal_area_mm2": 152.01,
            "diameter_m": 0.0139
        },
        {
            "MCM": 250,
            "AWG": None,
            "nominal_area_mm2": 126.68,
            "diameter_m": 0.0127
        },
        {
            "MCM": 211.6,
            "AWG": "4/0",
            "nominal_area_mm2": 107.22,
            "diameter_m": 0.0117
        },
        {
            "MCM": 167.8,
            "AWG": "3/0",
            "nominal_area_mm2": 85.03,
            "diameter_m": 0.0104
        },
        {
            "MCM": 133.1,
            "AWG": "2/0",
            "nominal_area_mm2": 67.44,
            "diameter_m": 0.0093
        },
        {
            "MCM": None,
            "AWG": "2",
            "nominal_area_mm2": 33.6,  # Valor corregido para AWG #2
            "diameter_m": 0.00654
        },
        {
            "MCM": None,
            "AWG": "4",
            "nominal_area_mm2": 21.1,  # Valor corregido para AWG #4
            "diameter_m": 0.00519
        }
    ]

    # Generar opciones para calibre_combo
    calibre_opciones = []
    for conductor in conductor_data:
        if conductor["AWG"] is not None:
            calibre_opciones.append(f"{conductor['AWG']} AWG")
        elif conductor["MCM"] is not None:
            calibre_opciones.append(f"{conductor['MCM']} MCM")

    tk.Label(frame_conductor, text="Calibre").grid(row=1, column=0, sticky="w", pady=5)
    calibre_combo = ttk.Combobox(frame_conductor, values=calibre_opciones, state="readonly")
    calibre_combo.grid(row=1, column=1, pady=5)
    calibre_combo.set(calibre_opciones[5])

    tk.Label(frame_conductor, text="Diámetro").grid(row=2, column=0, sticky="w", pady=5)
    diametro_label = tk.Label(frame_conductor, text="Dato Diámetro")
    diametro_label.grid(row=2, column=1, pady=5)

    tk.Label(frame_conductor, text="Sección").grid(row=3, column=0, sticky="w", pady=5)
    seccion_label = tk.Label(frame_conductor, text="Dato Sección")
    seccion_label.grid(row=3, column=1, pady=5)

    tk.Label(frame_conductor, text="Conductividad").grid(row=4, column=0, sticky="w", pady=5)
    conductividad_label = tk.Label(frame_conductor, text="Dato Conductividad")
    conductividad_label.grid(row=4, column=1, pady=5)

    tk.Label(frame_conductor, text="Resistividad (20°C)").grid(row=5, column=0, sticky="w", pady=5)
    resistividad_label = tk.Label(frame_conductor, text="Dato Resistividad")
    resistividad_label.grid(row=5, column=1, pady=5)

    # Función para verificar las distancias de los puntos
    def verificar_distancia_puntos():
        if not manager.archivo_cargado():
            messagebox.showwarning("Advertencia", "No hay ningún archivo cargado para realizar la validación.")
            return

        try:
            # Asegurarse de que la hoja "Área" existe
            if "Área" not in manager.workbook.sheetnames:
                messagebox.showerror("Error", "No se encontró la hoja 'Área' en el archivo Excel.")
                return

            # Acceder a la hoja "Área"
            area_sheet = manager.workbook["Área"]

            # Leer el tipo de geometría y los puntos
            geometry_type = area_sheet.cell(row=2, column=1).value
            puntos_str = area_sheet.cell(row=2, column=2).value
            if not puntos_str:
                messagebox.showerror("Error", "No hay puntos definidos en la hoja 'Área'.")
                return

            # Convertir los puntos de cadena a lista de tuplas
            puntos = eval(puntos_str)  # Evaluar la cadena para obtener la lista de puntos

            # Obtener la longitud de la varilla
            longitud_varilla = float(longitud_combo.get())
            distancia_minima = 2 * longitud_varilla  # Distancia mínima permitida entre puntos

            # Función para calcular la distancia entre dos puntos
            def calcular_distancia(p1, p2):
                return math.sqrt((p2[0] - p1[0]) ** 2 + (p2[1] - p1[1]) ** 2)

            # Decidir si cerrar los puntos dependiendo del tipo de geometría
            cerrar_puntos = geometry_type not in ["L", "Línea"]

            # Agregar el primer punto al final para cerrar, si corresponde
            if cerrar_puntos:
                puntos.append(puntos[0])

            # Comenzar a procesar los puntos y verificar las distancias
            nuevos_puntos = puntos[:]
            iteraciones = 0

            while True:
                se_agregaron_puntos = False
                nuevos_puntos_temp = []

                for i in range(len(nuevos_puntos) - 1):
                    p1 = nuevos_puntos[i]
                    p2 = nuevos_puntos[i + 1]
                    distancia = calcular_distancia(p1, p2)

                    nuevos_puntos_temp.append(p1)  # Agregar el punto actual

                    # Si la distancia no cumple la condición, agregar un punto intermedio
                    if distancia > distancia_minima:
                        punto_medio = ((p1[0] + p2[0]) / 2, (p1[1] + p2[1]) / 2)
                        nuevos_puntos_temp.append(punto_medio)
                        se_agregaron_puntos = True

                # Agregar el último punto
                nuevos_puntos_temp.append(nuevos_puntos[-1])

                # Actualizar la lista de puntos
                nuevos_puntos = nuevos_puntos_temp
                iteraciones += 1

                # Si no se agregaron más puntos, salir del bucle
                if not se_agregaron_puntos:
                    break

            # Mostrar los resultados en consola
            print(f"Tipo de geometría: {geometry_type}")
            print(f"Cantidad total de puntos: {len(nuevos_puntos)}")
            print("Puntos generados:")
            for idx, punto in enumerate(nuevos_puntos):
                print(f"Punto {idx + 1}: {punto}")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al procesar los puntos: {e}")

    # Ejecutar la verificación de distancias al abrir la ventana
    #verificar_distancia_puntos()

    # Función para cargar datos desde la hoja "Conductor"
    def cargar_datos_conductor():
        if not manager.archivo_cargado():
            return

        if "Conductor" not in manager.workbook.sheetnames:
            return

        try:
            conductor_sheet = manager.workbook["Conductor"]
            if conductor_sheet.max_row < 2:
                return  # No hay datos para cargar

            # Leer los datos de la fila 2 (o la primera fila con datos)
            fila_datos = 2
            datos = [conductor_sheet.cell(row=fila_datos, column=col).value for col in range(1, 10)]

            # Asignar valores a los combobox y etiquetas
            diametro_combo.set(datos[0])
            longitud_combo.set(datos[1])
            material_conductor_combo.set(datos[3])
            calibre_combo.set(datos[4])
            diametro_label.config(text=f"{datos[5]} m")
            seccion_label.config(text=f"{datos[6]} m²")
            conductividad_label.config(text=f"{datos[7]} %")
            resistividad_label.config(text=f"{datos[8]} μΩ∙cm")

        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron cargar los datos del conductor: {e}")

    def actualizar_datos_conductor(event=None):
        # Actualizar datos basados en material seleccionado
        selected_material = material_conductor_combo.get()
        material = next((m for m in Materiales_tipo if m["description"] == selected_material), None)
        if material:
            conductividad_label.config(text=f"{material['conductivity']} %")
            resistividad_label.config(text=f"{material['resistivity_20C']} μΩ∙cm")

        # Actualizar datos basados en conductor seleccionado
        selected_conductor = calibre_combo.get()
        conductor = next(
            (c for c in conductor_data if
             (f"{c['AWG']} AWG" if c["AWG"] else f"{c['MCM']} MCM") == selected_conductor), None)
        if conductor:
            diametro_label.config(text=f"{conductor['diameter_m']} m")
            seccion_label.config(text=f"{conductor['nominal_area_mm2']} m²")





    # Asociar la función al evento de cambio de selección en el combobox
    material_conductor_combo.bind("<<ComboboxSelected>>", actualizar_datos_conductor)
    calibre_combo.bind("<<ComboboxSelected>>", actualizar_datos_conductor)

    # Llamar a la función para actualizar los datos al abrir la ventana
    actualizar_datos_conductor()

    # Botones Guardar y Cancelar
    button_frame = tk.Frame(ventana_conductor)
    button_frame.pack(fill="x", pady=10)

    def guardar_datos():
        if not manager.archivo_cargado():
            messagebox.showwarning("Advertencia", "No hay ningún archivo abierto para guardar los datos.")
            return

        try:
            # Verificar si la hoja "Conductor" existe
            if "Conductor" not in manager.workbook.sheetnames:
                conductor_sheet = manager.workbook.create_sheet("Conductor")
                # Crear encabezados
                encabezados = [
                    "Diámetro Varilla", "Longitud Varilla (cm)", "Material Varilla",
                    "Material Conductor", "Calibre Conductor", "Diámetro Conductor",
                    "Sección Conductor", "Conductividad", "Resistividad (20°C)"
                ]
                for col, encabezado in enumerate(encabezados, start=1):
                    conductor_sheet.cell(row=1, column=col, value=encabezado)
            else:
                conductor_sheet = manager.workbook["Conductor"]

            # Buscar si ya hay datos en la hoja
            ultima_fila = conductor_sheet.max_row
            datos_existentes = False

            # Comprobar si hay datos a partir de la fila 2
            for row in conductor_sheet.iter_rows(min_row=2, max_row=ultima_fila, min_col=1, max_col=1,
                                                 values_only=True):
                if row[0]:  # Si hay datos en la primera columna (asume que siempre está llena)
                    datos_existentes = True
                    break

            # Establecer la fila donde se guardarán los datos
            fila_datos = 2 if datos_existentes else conductor_sheet.max_row + 1

            # Obtener los datos seleccionados
            diametro_varilla = diametro_combo.get()
            longitud_varilla = longitud_combo.get()
            material_conductor = material_conductor_combo.get()
            calibre_conductor = calibre_combo.get()

            # Encontrar los datos relacionados en Materiales_tipo
            material = next((m for m in Materiales_tipo if m["description"] == material_conductor), None)

            # Encontrar los datos relacionados en conductor_data
            conductor = next(
                (c for c in conductor_data if
                 (f"{c['AWG']} AWG" if c["AWG"] else f"{c['MCM']} MCM") == calibre_conductor), None)

            # Verificar que se hayan encontrado los datos
            if not material or not conductor:
                messagebox.showerror("Error", "No se pudo encontrar los datos relacionados para guardar.")
                return

            # Preparar los datos a guardar
            datos = [
                diametro_varilla,
                longitud_varilla,
                material_conductor,
                material["description"],
                calibre_conductor,
                conductor["diameter_m"],
                conductor["nominal_area_mm2"],
                material["conductivity"],
                material["resistivity_20C"],
            ]

            # Guardar los datos en la hoja
            for col, dato in enumerate(datos, start=1):
                conductor_sheet.cell(row=fila_datos, column=col, value=dato)

            # Guardar cambios en el archivo
            manager.guardar_archivo()

            # Mostrar mensaje de éxito
            messagebox.showinfo("Guardado", "Los datos del conductor se han guardado correctamente.")
            ventana_conductor.destroy()

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar los datos del conductor: {e}")

    def cancelar():
        ventana_conductor.destroy()

    tk.Button(button_frame, text="Guardar", command=guardar_datos).pack(padx=5)
    tk.Button(button_frame, text="Cancelar", command=cancelar).pack(padx=5)

    # Cargar datos al abrir la ventana
    cargar_datos_conductor()

puntos_rojos = []  # Lista para almacenar los puntos seleccionados como varillas

def abrir_diseno_malla_spt():
    # Crear la ventana para el diseño de malla SPT
    ventana_diseno = tk.Toplevel(root)
    ventana_diseno.title("Diseño de Malla SPT")
    ventana_diseno.geometry("720x450")
    ventana_diseno.resizable(True,True)

    # Canvas en la parte izquierda
    canvas_frame = tk.Frame(ventana_diseno)
    canvas_frame.grid(row=0, column=0, rowspan=2, padx=5, pady=5)
    canvas_width, canvas_height = 400, 400
    canvas = tk.Canvas(canvas_frame, width=canvas_width, height=canvas_height, bg="white")
    canvas.pack()

    # Frame de controles en la parte derecha
    control_frame = tk.Frame(ventana_diseno)
    control_frame.grid(row=0, column=1, padx=3, pady=10, sticky="n")

    # Sección de edición
    edicion_frame = tk.LabelFrame(control_frame, text="Edición", padx=5, pady=5)
    edicion_frame.pack(fill="x",  padx=10, pady=10)

    tk.Label(edicion_frame, text="N° de Varillas").grid(row=0, column=0, padx=5, sticky="w")
    varillas_combo = ttk.Combobox(edicion_frame, values=["Option 1", "Option 2"], state="readonly")
    varillas_combo.grid(row=0, column=1, padx=5, pady=5)

    tk.Label(edicion_frame, text="N° de Retículas").grid(row=1, column=0, padx=5, sticky="w")
    reticulas_combo = ttk.Combobox(edicion_frame, values=["1", "2"], state="readonly")
    reticulas_combo.grid(row=1, column=1, padx=5, pady=5)
    reticulas_combo.set("1")

    # Sección de variables
    variables_frame = tk.LabelFrame(control_frame, text="Variables", pady=5,padx=5)
    variables_frame.pack(fill="x", pady=10,padx=10)

    tk.Label(variables_frame, text="Longitud de conductor").grid(row=0, column=0, sticky="w",padx=2, pady=5)
    tk.Label(variables_frame, text="Longitud de varilla").grid(row=1, column=0, sticky="w",padx=2, pady=5)

    # Botones
    boton_frame = tk.Frame(control_frame)
    boton_frame.pack(fill="x", pady=5)

    global_scale_factor = 1  # Por defecto 1, cambiará en la función ajustar_y_dibujar_figura

    def ajustar_y_dibujar_figura(canvas, puntos, canvas_width, canvas_height,tipo_geometria):
        """
        Ajusta los puntos para que se adapten al tamaño del canvas y los dibuja,
        además calcula y muestra las distancias entre puntos originales.

        :param canvas: Canvas de Tkinter donde se dibuja.
        :param puntos: Lista de puntos [(x1, y1), (x2, y2), ...].
        :param canvas_width: Ancho del canvas.
        :param canvas_height: Altura del canvas.
        """
        global global_scale_factor  # Usaremos esta variable para guardar el valor actual de escala

        try:
            if not puntos or not all(isinstance(p, (list, tuple)) and len(p) == 2 for p in puntos):
                raise ValueError("Formato de puntos inválido. Se esperaba una lista de tuplas (x, y).")

            # Guardar los puntos originales para calcular las distancias
            puntos_originales = puntos[:]

            # Determinar escalado si los puntos exceden el tamaño del canvas
            max_x = max(p[0] for p in puntos)
            max_y = max(p[1] for p in puntos)
            scale_factor = min(canvas_width / max_x if max_x > canvas_width else 1,
                               canvas_height / max_y if max_y > canvas_height else 1)

            global_scale_factor = max(scale_factor - 0.1, 0.1)  # Reducir un poco el tamaño para evitar bordes

            # Escalar los puntos si es necesario
            if global_scale_factor < 1:
                puntos = [(x * global_scale_factor, y * global_scale_factor) for x, y in puntos]

            # Dibujar los puntos ajustados en el canvas
            canvas.delete("all")  # Limpiar el canvas antes de dibujar
            for i, punto in enumerate(puntos):
                x1, y1 = punto
                canvas.create_oval(x1 - 5, y1 - 5, x1 + 5, y1 + 5, fill="blue",
                                   tags="punto")  # Dibuja puntos como uniones

            print("Figura y distancias dibujadas con éxito.")
            actualizar_combobox_varillas(canvas, varillas_combo)
        except Exception as e:
            messagebox.showerror("Error", f"Error al dibujar figura: {e}")

    def cargar_datos_conductor():
        """
        Carga la longitud de las varillas desde la hoja 'Conductor' del archivo Excel.
        """
        if not manager.archivo_cargado():
            messagebox.showwarning("Advertencia", "No hay ningún archivo abierto para cargar los datos.")
            return None

        try:
            if "Conductor" not in manager.workbook.sheetnames:
                messagebox.showerror("Error", "No se encontró la hoja 'Conductor' en el archivo Excel.")
                return None

            # Leer la longitud de las varillas desde una celda específica
            conductor_sheet = manager.workbook["Conductor"]
            raw_value = conductor_sheet.cell(row=2, column=2).value  # Supongamos que está en esta celda
            print(raw_value)
            # Verifica que el valor no sea None y limpia si es necesario
            if raw_value is None:
                raise ValueError("El valor de longitud de varilla está vacío.")

            # Si el valor es una cadena, intenta extraer el número
            if isinstance(raw_value, str):
                import re
                match = re.search(r'\d+(\.\d+)?', raw_value)  # Busca un número decimal o entero
                if not match:
                    raise ValueError(f"El valor '{raw_value}' no contiene un número válido.")
                raw_value = float(match.group())

            # Verifica que sea numérico
            if not isinstance(raw_value, (int, float)):
                raise ValueError(f"El valor '{raw_value}' no es un número válido.")

            return raw_value

        except Exception as e:
            messagebox.showerror("Error", f"Hubo un problema al cargar los datos de 'Conductor': {e}")
            return None

    # Función para cargar y dibujar puntos desde la hoja "Área"
    def cargar_datos_y_dibujar():
        if not manager.archivo_cargado():
            messagebox.showwarning("Advertencia", "No hay ningún archivo abierto para cargar los datos.")
            return

        try:
            if "Área" not in manager.workbook.sheetnames:
                messagebox.showerror("Error", "No se encontró la hoja 'Área' en el archivo Excel.")
                return

            # Leer el tipo de geometría y los puntos desde la hoja "Área"
            area_sheet = manager.workbook["Área"]
            tipo_geometria = area_sheet.cell(row=2, column=1).value  # Supone que el tipo está en la columna 1
            puntos_str = area_sheet.cell(row=2, column=2).value      # Supone que los puntos están en la columna 2

            if not puntos_str:
                messagebox.showerror("Error", "No hay puntos definidos en la hoja 'Área'.")
                return

            puntos = eval(puntos_str)

            # Llamar a actualizar_reticulas después de cargar los puntos y tipo de geometría
            if tipo_geometria == "Rectángulo":
                actualizar_reticulas(tipo_geometria, puntos)

            # Verificar tipo de geometría y deshabilitar el combobox si es L o Línea
            if tipo_geometria in ["L", "Línea"]:
                tk.Label(edicion_frame, text="N° de Segmentos").grid(row=1, column=0, padx=5, sticky="w")
                actualizar_reticulas(tipo_geometria, puntos)

            # Ajustar y dibujar la figura
            ajustar_y_dibujar_figura(canvas, puntos, canvas_width, canvas_height, tipo_geometria)



        except Exception as e:
            messagebox.showerror("Error", f"Hubo un problema al cargar los datos: {e}")

    def actualizar_reticulas(tipo_geometria, puntos_originales):
        """
        Actualiza los valores del combobox de retículas según la geometría seleccionada.
        Calcula nuevos puntos si se cambian las divisiones.
        """
        if tipo_geometria == "L":
            # Permitir divisiones de 2 a 4
            reticulas_combo.config(values=["2", "3", "4"], state="readonly")
            reticulas_combo.set("2")  # Valor por defecto

            def dividir_y_actualizar(event):
                num_reticulas = int(reticulas_combo.get())
                nuevos_puntos = calcular_puntos_L(puntos_originales, num_reticulas)
                ajustar_y_dibujar_figura(canvas, nuevos_puntos, canvas_width, canvas_height, tipo_geometria)

        elif tipo_geometria == "Línea":
            reticulas_combo.config(values=[str(i) for i in range(2, 5)], state="readonly")
            reticulas_combo.set("2")  # Valor por defecto

            def dividir_y_actualizar(event):
                num_reticulas = int(reticulas_combo.get())
                nuevos_puntos = calcular_puntos_linea(puntos_originales, num_reticulas)  # Implementar para líneas
                ajustar_y_dibujar_figura(canvas, nuevos_puntos, canvas_width, canvas_height, tipo_geometria)

        elif tipo_geometria == "Rectángulo":
            reticulas_combo.config(values=["1", "2", "3", "4"], state="readonly")
            reticulas_combo.set("1")  # Valor por defecto

            def dividir_y_actualizar(event):
                num_reticulas = int(reticulas_combo.get())
                nuevos_puntos = calcular_puntos_reticulas(puntos_originales, num_reticulas, tipo_geometria)
                ajustar_y_dibujar_figura(canvas, nuevos_puntos, canvas_width, canvas_height, tipo_geometria)

        elif tipo_geometria == "Triángulo":
            reticulas_combo.config(values=["1", "2"], state="readonly")
            reticulas_combo.set("1")  # Valor por defecto

            def dividir_y_actualizar(event):
                num_reticulas = int(reticulas_combo.get())
                nuevos_puntos = calcular_puntos_triangulo(puntos_originales,
                                                          num_reticulas)  # Implementar para triángulos
                ajustar_y_dibujar_figura(canvas, nuevos_puntos, canvas_width, canvas_height, tipo_geometria)

        reticulas_combo.bind("<<ComboboxSelected>>", dividir_y_actualizar)

    def seleccionar_puntos_de_varillas():
        """
        Permite seleccionar puntos en el canvas para colocar varillas.
        """

        global puntos_rojos  # Declarar como global para usar la lista global

        # Obtener el número de varillas del combobox
        try:
            num_varillas = int(varillas_combo.get())
        except ValueError:
            messagebox.showerror("Error", "Por favor, seleccione un número válido de varillas.")
            return

        if num_varillas <= 0:
            messagebox.showerror("Error", "El número de varillas debe ser mayor que cero.")
            return

        # Obtener la longitud de la varilla desde Excel
        longitud_varilla = cargar_datos_conductor()
        if longitud_varilla is None:
            return  # Detener si no se pudo obtener el valor

        # Calcular la longitud total
        longitud_total = num_varillas * longitud_varilla

        # Mostrar la longitud total en el label
        V_longitud_varilla.config(text=f"{longitud_total} cm")


        # Deshabilitar los widgets
        varillas_combo.config(state="disabled")
        reticulas_combo.config(state="disabled")
        for widget in boton_frame.winfo_children():
            widget.config(state="disabled")

        # Mostrar mensaje de instrucciones
        messagebox.showinfo(
            "Seleccione Puntos de Varillas",
            "Seleccione los puntos en donde se colocarán las varillas."
        )

        # Lista para almacenar los puntos seleccionados
        puntos_canvas = [
            ((canvas.coords(p)[0] + canvas.coords(p)[2]) / 2, (canvas.coords(p)[1] + canvas.coords(p)[3]) / 2)
            for p in canvas.find_withtag("punto")
        ]

        puntos_seleccionados = []

        def seleccionar_punto(event):
            """
            Manejador de evento para seleccionar un punto en el canvas.
            Cambia el color del punto a rojo y guarda la selección.
            """
            nonlocal puntos_seleccionados
            global puntos_rojos

            # Coordenadas del clic
            x_click, y_click = event.x, event.y

            # Buscar el punto más cercano
            punto_cercano = None
            min_distancia = float('inf')
            for punto in puntos_canvas:
                x, y = punto
                distancia = ((x - x_click) ** 2 + (y - y_click) ** 2) ** 0.5
                if distancia < min_distancia and distancia <= 5:  # Tolerancia para seleccionar un punto
                    min_distancia = distancia
                    punto_cercano = punto

            if punto_cercano and punto_cercano not in puntos_seleccionados:
                # Cambiar el color del punto a rojo
                x, y = punto_cercano
                canvas.create_oval(x - 5, y - 5, x + 5, y + 5, fill="red")

                # Agregar a la lista de seleccionados
                puntos_seleccionados.append(punto_cercano)
                print("Puntos seleccionados",puntos_seleccionados)
                puntos_rojos.append(
                    (x / global_scale_factor, y / global_scale_factor))  # Ajustar según el factor de escala
                print("Puntos Rojos",puntos_rojos)

                # Verificar si ya se seleccionaron suficientes puntos
                if len(puntos_seleccionados) == num_varillas:
                    # Habilitar los widgets nuevamente
                    varillas_combo.config(state="readonly")
                    reticulas_combo.config(state="readonly")
                    for widget in boton_frame.winfo_children():
                        widget.config(state="normal")

                    # Desvincular el evento de clic
                    canvas.unbind("<Button-1>")

                    # Mensaje de confirmación
                    messagebox.showinfo(
                        "Selección Completa",
                        f"Se han seleccionado {num_varillas} puntos para las varillas."
                    )

        # Vincular el evento de clic
        canvas.bind("<Button-1>", seleccionar_punto)

    def calcular_puntos_L(puntos_originales, num_reticulas):
        """
        Calcula nuevos puntos para dividir una figura en forma de "L" en un número de retículas especificado.

        :param puntos_originales: Lista de puntos [(x1, y1), ...] que forman la "L".
        :param num_reticulas: Número de retículas deseadas.
        :return: Lista de puntos ajustados para las divisiones.
        """
        # Asumimos que los puntos originales están en el orden correcto para formar una "L":
        # P1 -> P2 -> P3 (donde P1-P2 es horizontal y P2-P3 es vertical).
        p1, p2, p3 = puntos_originales
        nuevos_puntos = puntos_originales[:]

        # Calcular longitudes de los segmentos
        dist_horizontal = ((p2[0] - p1[0]) ** 2 + (p2[1] - p1[1]) ** 2) ** 0.5
        dist_vertical = ((p3[0] - p2[0]) ** 2 + (p3[1] - p2[1]) ** 2) ** 0.5

        # Identificar el segmento más largo
        if dist_horizontal >= dist_vertical:
            segmento_largo = "horizontal"
            largo_inicio, largo_fin = p1, p2
            corto_inicio, corto_fin = p2, p3
        else:
            segmento_largo = "vertical"
            largo_inicio, largo_fin = p2, p3
            corto_inicio, corto_fin = p1, p2

        # Calcular puntos según el número de retículas
        if num_reticulas == 2:
            # Agregar un punto en la mitad del segmento más largo
            t = 0.5
            nuevo_punto = (
                (1 - t) * largo_inicio[0] + t * largo_fin[0],
                (1 - t) * largo_inicio[1] + t * largo_fin[1],
            )
            nuevos_puntos.append(nuevo_punto)

        elif num_reticulas == 3:
            # Dividir el segmento más largo en dos partes
            t1, t2 = 1 / 3, 2 / 3
            punto1 = (
                (1 - t1) * largo_inicio[0] + t1 * largo_fin[0],
                (1 - t1) * largo_inicio[1] + t1 * largo_fin[1],
            )
            punto2 = (
                (1 - t2) * largo_inicio[0] + t2 * largo_fin[0],
                (1 - t2) * largo_inicio[1] + t2 * largo_fin[1],
            )
            nuevos_puntos.extend([punto1, punto2])


        elif num_reticulas == 4:

            # Dividir el segmento horizontal en 2 partes

            for i in range(1, 2):  # Divisiones internas (sin extremos)

                t_horizontal = i / 2

                punto_horizontal = (

                    (1 - t_horizontal) * corto_inicio[0] + t_horizontal * corto_fin[0],

                    (1 - t_horizontal) * corto_inicio[1] + t_horizontal * corto_fin[1],

                )

                nuevos_puntos.append(punto_horizontal)

            # Dividir el segmento vertical en 3 partes

            for i in range(1, 3):  # Divisiones internas (sin extremos)

                t_vertical = i / 3

                punto_vertical = (

                    (1 - t_vertical) * largo_inicio[0] + t_vertical * largo_fin[0],

                    (1 - t_vertical) * largo_inicio[1] + t_vertical * largo_fin[1],

                )

                nuevos_puntos.append(punto_vertical)
        print(nuevos_puntos)
        # Asegurar que los puntos estén en el orden correcto para dibujar
        return sorted(nuevos_puntos, key=lambda p: (p[1], p[0]))

    def calcular_puntos_linea(puntos_originales, num_reticulas):
        """
        Calcula nuevos puntos para dividir una línea en segmentos iguales.

        :param puntos_originales: Lista de puntos [(x1, y1), (x2, y2)].
        :param num_reticulas: Número de segmentos.
        :return: Lista de puntos ajustados para las divisiones.
        """
        if len(puntos_originales) != 2:
            raise ValueError("Se requieren exactamente 2 puntos para definir una línea.")

        x1, y1 = puntos_originales[0]
        x2, y2 = puntos_originales[1]
        nuevos_puntos = [puntos_originales[0]]  # Agregar el primer punto

        for i in range(1, num_reticulas):
            t = i / num_reticulas
            nuevo_punto = ((1 - t) * x1 + t * x2, (1 - t) * y1 + t * y2)
            nuevos_puntos.append(nuevo_punto)

        nuevos_puntos.append(puntos_originales[1])  # Agregar el último punto
        return nuevos_puntos

    def calcular_puntos_reticulas(puntos_originales, num_reticulas, tipo_geometria):
        """
        Calcula nuevos puntos para dividir un rectángulo en un número de retículas especificado.

        :param puntos_originales: Lista de puntos [(x1, y1), (x2, y2), ...].
        :param num_reticulas: Número de retículas (horizontal y vertical).
        :param tipo_geometria: Tipo de geometría (para futuros ajustes).
        :return: Lista de puntos ajustados para las divisiones.
        """
        if len(puntos_originales) != 4:
            raise ValueError("Se requieren exactamente 4 puntos para definir un rectángulo.")

        x1, y1 = puntos_originales[0]
        x2, y2 = puntos_originales[1]
        x3, y3 = puntos_originales[2]
        x4, y4 = puntos_originales[3]

        nuevos_puntos = []

        # Divisiones horizontales
        for i in range(num_reticulas + 1):
            t = i / num_reticulas
            izquierda_x = (1 - t) * x1 + t * x4
            izquierda_y = (1 - t) * y1 + t * y4
            derecha_x = (1 - t) * x2 + t * x3
            derecha_y = (1 - t) * y2 + t * y3

            for j in range(num_reticulas + 1):
                t_horiz = j / num_reticulas
                nuevo_x = (1 - t_horiz) * izquierda_x + t_horiz * derecha_x
                nuevo_y = (1 - t_horiz) * izquierda_y + t_horiz * derecha_y
                nuevos_puntos.append((nuevo_x, nuevo_y))

        return sorted(set(nuevos_puntos), key=lambda p: (p[1], p[0]))

    def actualizar_combobox_varillas(canvas, varillas_combo):
        """
        Actualiza las opciones del combobox 'Nº de Varillas' según la cantidad de puntos en el canvas.

        :param canvas: Canvas de Tkinter donde se dibujan los puntos.
        :param varillas_combo: Combobox de Tkinter para 'Nº de Varillas'.
        """
        # Contar la cantidad de puntos en el canvas
        cantidad_puntos = len(canvas.find_withtag("punto"))

        if cantidad_puntos > 0:
            # Actualizar las opciones del combobox con números desde 1 hasta la cantidad de puntos
            opciones = [str(i) for i in range(1, cantidad_puntos + 1)]
            varillas_combo.config(values=opciones)
            varillas_combo.set("1")  # Establecer el valor por defecto en "1"
        else:
            # Si no hay puntos, deshabilitar el combobox
            varillas_combo.config(values=["1"], state="disabled")

    def conectar_puntos():
        """
        Permite conectar puntos en el canvas mediante líneas.
        """
        global global_scale_factor  # Usamos este valor para desescalar las distancias
        global puntos_conectados
        global distancia_total  # Nueva variable para almacenar la distancia total
        global lineas_creadas  # Declarar como global para ser accesible


        distancia_total = 0  # Inicializar la distancia total

        messagebox.showinfo(
            "Conectar Puntos",
            "Seleccione dos puntos para unirlos. Al terminar pulse el botón 'Ok'."
        )

        # Deshabilitar widgets
        varillas_combo.config(state="disabled")
        reticulas_combo.config(state="disabled")
        for widget in boton_frame.winfo_children():
            widget.config(state="disabled")

        # Crear botón temporal de "Ok"
        boton_ok = tk.Button(canvas, text="Ok", bg="lightgray")
        boton_ok.place(x=10, y=10)

        boton_deshacer = tk.Button(canvas, text="Deshacer", bg="lightgray")
        boton_deshacer.place(x=60, y=10)

        puntos_canvas = [
            (
                (canvas.coords(p)[0] + canvas.coords(p)[2]) / 2,
                (canvas.coords(p)[1] + canvas.coords(p)[3]) / 2,
            )
            for p in canvas.find_withtag("punto")
        ]

        puntos_seleccionados = []
        lineas_creadas = []  # Almacena las líneas creadas y sus textos asociados

        def calcular_dibujar_distancia(canvas, x1, y1, x2, y2, global_scale_factor):
            """
            Calcula la distancia entre dos puntos y la muestra en el canvas.
            La distancia se coloca de manera vertical, horizontal o diagonal según la orientación de la línea.

            :param canvas: Canvas de Tkinter donde se dibuja la línea.
            :param x1: Coordenada x del primer punto.
            :param y1: Coordenada y del primer punto.
            :param x2: Coordenada x del segundo punto.
            :param y2: Coordenada y del segundo punto.
            :param global_scale_factor: Escala aplicada al canvas para desescalar las distancias.
            """

            global distancia_total

            # Calcular la distancia original (sin escala)
            distancia = ((x2 / global_scale_factor - x1 / global_scale_factor) ** 2 +
                         (y2 / global_scale_factor - y1 / global_scale_factor) ** 2) ** 0.5
            distancia_cm = round(distancia, 2)

            # Acumular la distancia total
            distancia_total += distancia_cm

            # Dibujar la línea con ancho 2
            linea = canvas.create_line(x1, y1, x2, y2, width=2)

            # Determinar orientación de la línea y colocar el texto
            if abs(x2 - x1) > abs(y2 - y1):  # Línea horizontal
                x_text = (x1 + x2) / 2  # Centro horizontal
                y_text = max(y1, y2) + 10  # Debajo de la línea
                texto = canvas.create_text(x_text, y_text, text=f"{distancia_cm} cm", fill="black", font=("Arial", 8))
            elif abs(y2 - y1) > abs(x2 - x1):  # Línea vertical
                x_text = min(x1, x2) - 10  # A la izquierda de la línea
                y_text = (y1 + y2) / 2  # Centro vertical
                texto = canvas.create_text(x_text, y_text, text=f"{distancia_cm} cm", fill="black", font=("Arial", 8),
                                           angle=90)
            else:  # Línea diagonal
                x_text = (x1 + x2) / 2  # Centro de la línea
                y_text = (y1 + y2) / 2  # Centro de la línea
                texto = canvas.create_text(x_text, y_text, text=f"{distancia_cm} cm", fill="black", font=("Arial", 8),
                                           angle=-45)

            # Guardar la línea y el texto para permitir "Deshacer"
            lineas_creadas.append((linea, texto))

        def seleccionar_punto(event):
            """
            Manejador de evento para seleccionar un punto en el canvas y conectar dos puntos.
            """
            nonlocal puntos_seleccionados
            global puntos_conectados

            # Coordenadas del clic
            x_click, y_click = event.x, event.y

            # Buscar el punto más cercano al clic
            punto_cercano = None
            min_distancia = float('inf')
            for punto in puntos_canvas:
                x, y = punto
                distancia = ((x - x_click) ** 2 + (y - y_click) ** 2) ** 0.5
                if distancia < min_distancia and distancia <= 10:  # Ajusta esta sensibilidad si es necesario
                    min_distancia = distancia
                    punto_cercano = punto

            if punto_cercano:
                # Cambiar el color del punto a amarillo
                x, y = punto_cercano
                canvas.create_oval(x - 5, y - 5, x + 5, y + 5, fill="yellow", tags="selected")

                # Agregar a la lista de seleccionados
                puntos_seleccionados.append(punto_cercano)

                # Si se seleccionaron dos puntos, dibujar la línea
                if len(puntos_seleccionados) == 2:
                    x1, y1 = puntos_seleccionados[0]
                    x2, y2 = puntos_seleccionados[1]

                    # Dibujar la línea y calcular la distancia
                    calcular_dibujar_distancia(canvas, x1, y1, x2, y2, global_scale_factor)

                    # Cambiar el estado de puntos_conectados
                    puntos_conectados = True

                    # Restaurar los puntos seleccionados a azul
                    for px, py in puntos_seleccionados:
                        canvas.create_oval(px - 5, py - 5, px + 5, py + 5, fill="blue", tags="punto")

                    puntos_seleccionados.clear()

        def deshacer_ultima_linea():
            """
            Deshace la última línea creada junto con su texto asociado.
            """
            if lineas_creadas:
                linea, texto = lineas_creadas.pop()
                canvas.delete(linea)
                canvas.delete(texto)

                # Actualizar el estado de puntos_conectados
                global puntos_conectados
                puntos_conectados = len(lineas_creadas) > 0



        def finalizar_conexion():
            """
            Finaliza la conexión de puntos, elimina el botón "Ok", y habilita los widgets.
            """
            canvas.unbind("<Button-1>")

            boton_ok.destroy()
            boton_deshacer.destroy()

            varillas_combo.config(state="readonly")
            reticulas_combo.config(state="readonly")
            for widget in boton_frame.winfo_children():
                widget.config(state="normal")

            if puntos_conectados:
                boton_seleccionar_varillas.config(state="normal")  # Habilitar si hay conexiones
            else:
                boton_seleccionar_varillas.config(state="disabled")  # Mantener deshabilitado si no hay conexiones

            print("Distancia Total:", distancia_total)

            boton_conectar_p.config(state="disabled")
            reticulas_combo.config(state="disabled")
            V_longitud_conductor.config(text=f"{distancia_total} cm")


        boton_ok.config(command=finalizar_conexion)
        boton_deshacer.config(command=deshacer_ultima_linea)

        canvas.bind("<Button-1>", seleccionar_punto)

    def guardar_datos():
        """
        Guarda los datos del diseño de malla en la hoja "Diseño Malla" del archivo Excel.
        Sobrescribe los datos si ya existen.
        """
        global puntos_rojos, distancia_total, lineas_creadas, global_scale_factor

        if not manager.archivo_cargado():
            messagebox.showwarning("Advertencia", "No hay ningún archivo cargado para guardar los datos.")
            return

        try:
            # Asegurarse de que exista la hoja "Diseño Malla"
            if "Diseño Malla" not in manager.workbook.sheetnames:
                manager.workbook.create_sheet("Diseño Malla")

            # Seleccionar la hoja "Diseño Malla"
            diseño_malla_sheet = manager.workbook["Diseño Malla"]

            # Encabezados (si están vacíos)
            encabezados = ["Puntos Rojos", "Longitud Total Conductor", "N° de Varillas",
                           "Longitud Total Varillas", "Líneas (coordenadas ajustadas)"]
            for col, header in enumerate(encabezados, start=1):
                if diseño_malla_sheet.cell(row=1, column=col).value is None:
                    diseño_malla_sheet.cell(row=1, column=col).value = header

            # Ajustar los puntos rojos por el inverso del factor de escala
            puntos_rojos_ajustados = [(x / global_scale_factor, y / global_scale_factor) for x, y in puntos_rojos]

            # Guardar los datos en la segunda fila
            diseño_malla_sheet.cell(row=2, column=1).value = str(
                puntos_rojos_ajustados)  # Guardar puntos rojos ajustados
            diseño_malla_sheet.cell(row=2, column=2).value = distancia_total  # Guardar distancia total
            diseño_malla_sheet.cell(row=2, column=3).value = varillas_combo.get()  # Número de varillas
            longitud_varilla = cargar_datos_conductor()
            if longitud_varilla is not None:
                diseño_malla_sheet.cell(row=2, column=4).value = int(varillas_combo.get()) * longitud_varilla

            # Guardar las líneas creadas ajustadas por el factor de escala
            líneas_ajustadas = [
                ((canvas.coords(linea[0])[0] / global_scale_factor,
                  canvas.coords(linea[0])[1] / global_scale_factor),
                 (canvas.coords(linea[0])[2] / global_scale_factor,
                  canvas.coords(linea[0])[3] / global_scale_factor))
                for linea in lineas_creadas
            ]
            diseño_malla_sheet.cell(row=2, column=5).value = str(líneas_ajustadas)

            # Guardar el archivo
            manager.guardar_archivo()

            # Confirmación de guardado
            messagebox.showinfo("Guardado", "Datos guardados correctamente en 'Diseño Malla'.")

        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron guardar los datos: {e}")

    #Declaración de labels para mostrar datos
    V_longitud_conductor = tk.Label(variables_frame, text="Valor")
    V_longitud_conductor.grid(row=0, column=1, sticky="e",padx=5, pady=5)
    V_longitud_varilla = tk.Label(variables_frame, text="Valor")
    V_longitud_varilla.grid(row=1, column=1, sticky="e",padx=5, pady=5)

    # Llamar a la función para cargar y dibujar al abrir la ventana
    cargar_datos_y_dibujar()
    boton_conectar_p = tk.Button(boton_frame, text="Conectar Puntos", width=25, command=conectar_puntos)
    boton_conectar_p.pack(pady=5)
    boton_seleccionar_varillas = tk.Button(boton_frame, text="Seleccionar Varillas", width=25, command=seleccionar_puntos_de_varillas, state="disabled")
    boton_seleccionar_varillas.pack(pady=5)
    tk.Button(boton_frame, text="Guardar", width=25, command=guardar_datos).pack(pady=5)
    tk.Button(boton_frame, text="Cancelar", width=25, command=ventana_diseno.destroy).pack(pady=5)


def abrir_analisis_cortocircuito():
    """
    Abre una nueva ventana para realizar el análisis de cortocircuito.
    """
    # Crear la nueva ventana
    ventana_cortocircuito = tk.Toplevel(root)
    ventana_cortocircuito.title("Análisis de Cortocircuito")
    ventana_cortocircuito.geometry("400x300")
    ventana_cortocircuito.resizable(False, False)

    # Contenedor principal
    frame_principal = tk.Frame(ventana_cortocircuito)
    frame_principal.pack(fill="both", expand=True, padx=20, pady=20)

    # Campos de entrada para ICC
    tk.Label(frame_principal, text="Corriente simétrica de falla").grid(row=0, column=0, sticky="e", padx=5, pady=10)
    corriente_sim_falla = tk.Entry(frame_principal, width=15)
    corriente_sim_falla.grid(row=0, column=1, padx=5, pady=10)

    tk.Label(frame_principal, text="Corriente máxima de falla").grid(row=1, column=0, sticky="e", padx=5, pady=10)
    corriente_max_falla = tk.Entry(frame_principal, width=15)
    corriente_max_falla.grid(row=1, column=1, padx=5, pady=10)

    # Combobox para seleccionar tf
    tk.Label(frame_principal, text="Tiempo de Falla").grid(row=2, column=0, sticky="e", padx=5, pady=10)
    tf_var = tk.StringVar(value="0.05")  # Valor por defecto para tf
    tf_combobox = ttk.Combobox(frame_principal, textvariable=tf_var,
                               values=["0.00833", "0.05", "0.10", "0.20", "0.30", "0.40", "0.50", "0.75", "1.00"],
                               state="readonly")  # Opciones y estado de solo lectura
    tf_combobox.grid(row=2, column=1, padx=5, pady=10)

    # Combobox para seleccionar X/R
    tk.Label(frame_principal, text="X/R").grid(row=3, column=0, sticky="e", padx=5, pady=10)
    xr_var = tk.StringVar(value="10")  # Valor por defecto para X/R
    xr_combobox = ttk.Combobox(frame_principal, textvariable=xr_var,
                               values=["10", "20", "30", "40"],
                               state="readonly")  # Opciones y estado de solo lectura
    xr_combobox.grid(row=3, column=1, padx=5, pady=10)

    # Diccionario de datos
    data = {
        0.00833: {10: 1.576, 20: 1.648, 30: 1.675, 40: 1.688},
        0.05: {10: 1.232, 20: 1.378, 30: 1.462, 40: 1.515},
        0.10: {10: 1.125, 20: 1.232, 30: 1.316, 40: 1.378},
        0.20: {10: 1.064, 20: 1.125, 30: 1.181, 40: 1.232},
        0.30: {10: 1.043, 20: 1.085, 30: 1.125, 40: 1.163},
        0.40: {10: 1.033, 20: 1.064, 30: 1.095, 40: 1.125},
        0.50: {10: 1.026, 20: 1.052, 30: 1.077, 40: 1.101},
        0.75: {10: 1.018, 20: 1.035, 30: 1.052, 40: 1.070},
        1.00: {10: 1.013, 20: 1.026, 30: 1.039, 40: 1.052},
    }

    # Botón de guardar
    def guardar_datos_cortocircuito():
        print("Botón Guardar")

    # Botón de calcular
    def Calcular_cortocircuito():
        try:
            # Obtener los valores seleccionados de los combobox
            tf = float(tf_var.get())
            xr = int(xr_var.get())

            # Cargar los datos del Excel
            if not manager.archivo_cargado():
                tk.messagebox.showerror("Error", "No hay ningún archivo Excel cargado.")
                return

            # Verificar si la hoja "trafos" existe
            if "trafos" not in manager.workbook.sheetnames:
                tk.messagebox.showerror("Error", "No se encontró la hoja 'trafos' en el archivo Excel.")
                return

            # Acceder a la hoja "trafos"
            trafos_sheet = manager.workbook["trafos"]

            # Obtener los datos requeridos
            try:
                ultima_fila = trafos_sheet.max_row  # Suponiendo que los datos están en la última fila
                KVA = float(trafos_sheet.cell(row=ultima_fila, column=3).value or 0)  # Potencia nominal (kVA)
                EMt = float(trafos_sheet.cell(row=ultima_fila, column=1).value or 0)  # Tensión del primario (V)
                EBt = float(trafos_sheet.cell(row=ultima_fila, column=2).value or 0)  # Tensión del secundario (V)
                Z = float(trafos_sheet.cell(row=ultima_fila, column=6).value.split()[0])  # Impedancia (%)
                Sf=0.20;
            except ValueError as e:
                tk.messagebox.showerror("Error", f"No se pudo obtener datos válidos del Excel: {e}")
                return
            IsecBT = (KVA * 1000) / (math.sqrt(3) * EBt)
            IsecMT = (KVA * 1000) / (math.sqrt(3) * EMt)
            ICCmaxBT = (100 / Z) * IsecBT
            ICCmaxMT = (100 / Z) * IsecMT

            # Validar que los valores tf y xr existen en el diccionario
            if tf in data and xr in data[tf]:
                Df = data[tf][xr]
                Icc_asimBT = ICCmaxBT* Df
                Icc_asimMT = ICCmaxMT * Df
                Ig= Sf * Icc_asimBT #Corriente simétrica de falla
                IG= Df * Ig         #Corriente máxima de falla


                # Mostrar los resultados en los campos
                corriente_sim_falla.delete(0, tk.END)  # Limpiar campo
                corriente_sim_falla.insert(0, f"{Ig:.2f}")  # Insertar resultado de Ig

                corriente_max_falla.delete(0, tk.END)  # Limpiar campo
                corriente_max_falla.insert(0, f"{IG:.2f}")  # Insertar resultado de IG


            else:
                tk.messagebox.showerror("Error", "La combinación seleccionada no es válida.")

        except Exception as e:
            tk.messagebox.showerror("Error", f"Ha ocurrido un error: {e}")




    tk.Button(frame_principal, text="Guardar", command=guardar_datos_cortocircuito).grid(row=4, column=1, columnspan=2, pady=20)
    tk.Button(frame_principal, text="Calcular", command=Calcular_cortocircuito).grid(row=4, column=0, columnspan=2, pady=20)



def abrir_analisis_resistividad():
    """
    Abre una ventana para realizar el análisis de resistividad,
    tomando datos directamente de la hoja 'Resistencias' del archivo Excel.
    """
    if not manager.archivo_cargado():
        messagebox.showwarning("Advertencia", "No hay ningún archivo abierto para realizar el análisis.")
        return

    # Crear la nueva ventana
    ventana_analisis = tk.Toplevel(root)
    ventana_analisis.title("Análisis de Resistividad")
    ventana_analisis.geometry("800x600")

    # Crear un frame para el gráfico
    frame_grafico = tk.Frame(ventana_analisis)
    frame_grafico.pack(fill="both", expand=True, padx=10, pady=10)

    try:
        if "Resistencias" not in manager.workbook.sheetnames:
            messagebox.showerror("Error", "No se encontró la hoja 'Resistencias' en el archivo Excel.")
            return

        # Acceder a la hoja 'Resistencias'
        resistencias_sheet = manager.workbook["Resistencias"]

        # Leer los encabezados de la hoja para obtener las distancias
        encabezados = [cell.value for cell in resistencias_sheet[1] if cell.value]
        if "Perfil" not in encabezados:
            messagebox.showerror("Error", "La hoja 'Resistencias' no tiene un encabezado válido.")
            return

        distancias = [
            float(etiq.replace('m', ''))  # Extraer la distancia numérica de etiquetas como "2m"
            for etiq in encabezados[1:] if etiq and etiq.endswith("m")
        ]


        # Leer los datos de resistencias desde la hoja
        resistencias = []
        etiquetas_perfil = []

        for row in resistencias_sheet.iter_rows(min_row=2, values_only=True):
            perfil, *valores_resistencias = row
            if perfil:  # Solo incluir filas con un perfil
                etiquetas_perfil.append(perfil)
                resistencias.append([
                    float(valor) if valor not in [None, ''] and str(valor).replace('.', '', 1).isdigit() else None
                    for valor in valores_resistencias
                ])

        # Verificar si hay datos válidos
        if not resistencias or not distancias:
            messagebox.showerror("Error", "No se encontraron datos válidos en la hoja 'Resistencias'.")
            return

        # Calcular resistividades
        resistividades = []
        for fila_res in resistencias:
            resistividades.append([
                2 * math.pi * distancias[j] * r
                for j, r in enumerate(fila_res) if r is not None
            ])

        # Calcular promedio de resistividades
        promedio_resistividades = []
        for i in range(len(distancias)):
            resistividades_columna = [fila[i] for fila in resistividades if i < len(fila)]
            promedio_resistividades.append(
                sum(resistividades_columna) / len(resistividades_columna) if resistividades_columna else None
            )

        # Crear el gráfico
        fig, ax = plt.subplots(figsize=(8, 6))

        # Graficar resistividades de cada perfil
        colores = ['g', 'orange', 'purple', 'blue']
        for i, fila_resistividades in enumerate(resistividades):
            ax.plot(distancias, fila_resistividades, label=etiquetas_perfil[i],
                    marker='o', color=colores[i % len(colores)])

        # Graficar línea de promedio
        promedio_resistividades = [val for val in promedio_resistividades if val is not None]
        ax.plot(distancias, promedio_resistividades, label="Promedio",
                marker='x', color='red', linestyle='--', linewidth=2)

        # Configuración del gráfico
        ax.set_title("GRÁFICO DE RESISTIVIDADES", fontsize=14, fontweight='bold')
        ax.set_xlabel("SEPARACIÓN ENTRE ELECTRODOS (m)", fontsize=12)
        ax.set_ylabel("RESISTIVIDAD ($\Omega \cdot m$)", fontsize=12)
        ax.legend(loc='upper right', fontsize=10)
        ax.grid(True, linestyle='--', alpha=0.7)

        # Mostrar el gráfico en la ventana Tkinter
        canvas = FigureCanvasTkAgg(fig, master=frame_grafico)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo realizar el análisis de resistividad: {e}")


def calcular_resistividad_con_profundidad():
    """
    Calcula resistividades, promedios por distancia y solicita profundidad.
    Luego obtiene Lc, dc, A y r desde hojas específicas, realizando conversiones.
    """
    if not manager.archivo_cargado():
        print("No hay ningún archivo abierto para calcular resistividades.")
        return

    try:
        # Verificar si la hoja "Resistencias" existe
        if "Resistencias" not in manager.workbook.sheetnames:
            print("No se encontró la hoja 'Resistencias' en el archivo Excel.")
            return

        # Acceder a la hoja "Resistencias"
        resistencias_sheet = manager.workbook["Resistencias"]

        # Obtener los encabezados de la primera fila
        encabezados = [cell.value for cell in resistencias_sheet[1] if cell.value]
        print("Encabezados encontrados:", encabezados)

        # Filtrar columnas válidas (distancias como "2m", "3m", etc.)
        columnas_validas = [col for col in encabezados if "m" in str(col).lower()]
        print("Columnas válidas detectadas:", columnas_validas)

        # Crear un diccionario para almacenar las resistividades de cada columna
        resistividades = {col: [] for col in columnas_validas}

        # Recorrer las filas desde la segunda fila
        for row in resistencias_sheet.iter_rows(min_row=2, max_row=resistencias_sheet.max_row, values_only=True):
            fila_dict = dict(zip(encabezados, row))  # Convertir la fila en un diccionario clave:valor
            for col in columnas_validas:
                valor = fila_dict.get(col)  # Acceder al valor por nombre de columna
                try:
                    # Extraer la distancia (por ejemplo, "2m" → 2)
                    distancia = float(col.lower().replace('m', '').strip())

                    # Convertir el valor a número si es una cadena
                    if isinstance(valor, str):
                        valor = float(valor.replace(',', '.'))

                    if isinstance(valor, (int, float)) and valor > 0:  # Verificar que sea numérico y positivo
                        # Calcular resistividad usando la fórmula
                        resistividad = 2 * math.pi * distancia * valor
                        resistividades[col].append(resistividad)
                except (ValueError, TypeError):
                    continue

        # Calcular los promedios de resistividades
        promedios = {}
        for col, valores in resistividades.items():
            promedios[col] = sum(valores) / len(valores) if valores else None

        # Mostrar promedios
        print("\n--- Promedio de Resistividad por Distancia ---")
        for col, promedio in promedios.items():
            if promedio is not None:
                print(f"{col}: {promedio:.2f} Ω·m")
            else:
                print(f"{col}: No hay datos disponibles")

        # Solicitar distancia seleccionada
        print("\nPor favor, elija una distancia de las disponibles:", ", ".join(columnas_validas))
        medida_elegida = input("Distancia seleccionada: ").strip()

        if medida_elegida not in promedios or promedios[medida_elegida] is None:
            print("La distancia seleccionada no es válida o no tiene datos.")
            return
        p=promedios[medida_elegida]
        print(f"Usted seleccionó '{medida_elegida}' con promedio: {promedios[medida_elegida]:.2f} Ω·m")

        # Solicitar valor de profundidad h
        h = float(input("Ingrese el valor de profundidad de los conductores (h) en metros: ").strip())

        # Obtener "Longitud total de conductor" (Lc) desde la hoja 'Diseño Malla'
        if "Diseño Malla" not in manager.workbook.sheetnames:
            print("No se encontró la hoja 'Diseño Malla'.")
            return

        diseño_malla_sheet = manager.workbook["Diseño Malla"]
        Lc = diseño_malla_sheet.cell(row=2, column=2).value  # Columna 2: Longitud total de conductor
        if Lc is None:
            print("No se encontró un valor para la Longitud total de conductor (Lc).")
            return
        Lc_metros = float(Lc) / 100  # Convertir a metros
        print(f"Longitud total del conductor (Lc): {Lc_metros:.2f} m")

        Lv = diseño_malla_sheet.cell(row=2, column=4).value  # Columna 4: Longitud total de varillas
        if Lv is None:
            print("No se encontró un valor para la Longitud total de varillas (Lv).")
            return
        Lv_metros = float(Lv) / 100  # Convertir a metros
        print(f"Longitud total del varillas (Lv): {Lv_metros:.2f} m")

        Nv = diseño_malla_sheet.cell(row=2, column=3).value  # Columna 3: Número de Varillas
        if Nv is None:
            print("No se encontró un valor para la Longitud total de varillas (Lv).")
            return
        Nv=float(Nv)
        print(f"Número de varillas: {Nv}")

        # Obtener "Diámetro del conductor" (dc) desde la hoja 'Conductor'
        if "Conductor" not in manager.workbook.sheetnames:
            print("No se encontró la hoja 'Conductor'.")
            return

        conductor_sheet = manager.workbook["Conductor"]
        dc = conductor_sheet.cell(row=2, column=6).value  # Columna 6: Diámetro del conductor
        if dc is None:
            print("No se encontró un valor para el Diámetro del conductor (dc).")
            return
        dc = float(dc)
        print(f"Diámetro del conductor (dc): {dc:.4f} m")

        # Obtener valores de Área (A) y Radio equivalente (r) desde la hoja 'Área'
        if "Área" not in manager.workbook.sheetnames:
            print("No se encontró la hoja 'Área'.")
            return

        area_sheet = manager.workbook["Área"]
        a = area_sheet.cell(row=2, column=5).value  # Columna 5: Área
        a = float(a)
        print("a=",a)
        r = area_sheet.cell(row=2, column=6).value  # Columna 6: Radio equivalente
        r = float(r)  # Convertir a flotante
        print("r=", r)

        lt=float(Lc_metros+(Nv*Lv_metros))
        print("Lt=",lt)
        rg=float(p*((1/lt)+(1/math.sqrt(20*a))*(1+(1/1+h*(math.sqrt(20/a))))))
        print("Rg=", rg)

        if a is None or r is None:
            print("No se encontraron valores para Área (A) o Radio equivalente (r).")
            return

        print(f"Área (A): {a:.4f} m²")
        print(f"Radio equivalente (r): {r:.4f} m")

        # Mostrar todos los valores obtenidos
        print("\n--- Valores Obtenidos ---")
        print(f"Distancia seleccionada: {medida_elegida}")
        print(f"Promedio de Resistividad: {promedios[medida_elegida]:.2f} Ω·m")
        print(f"Profundidad (h): {h:.2f} m")
        print(f"Longitud total del conductor (Lc): {Lc_metros:.2f} m")
        print(f"Diámetro del conductor (dc): {dc:.4f} m")
        print(f"Área (A): {a:.2f} m²")
        print(f"Radio equivalente (r): {r:.4f} m")
        print(f"Resistencia total de la malla (Rg): {rg:.4f} Ω")
    except Exception as e:
        print(f"Error al calcular resistividades y obtener valores adicionales: {e}")








tk.Button(boton_Calcular_frame, text="Cálculo", command=calcular_resistividad_con_profundidad).grid(row=0, column=0)

def deshabilitar_menu_edit():
    edit_menu.entryconfig("Datos de proyecto", state="disabled")
    edit_menu.entryconfig("Datos de resistencia", state="disabled")
    edit_menu.entryconfig("Datos de transformador", state="disabled")
    edit_menu.entryconfig("Datos de conductor", state="disabled")
    edit_menu.entryconfig("Área para el SPT", state="disabled")


def habilitar_menu_edit():
    edit_menu.entryconfig("Datos de proyecto", state="normal")
    edit_menu.entryconfig("Datos de resistencia", state="normal")
    edit_menu.entryconfig("Datos de transformador", state="normal")
    edit_menu.entryconfig("Datos de conductor", state="normal")
    edit_menu.entryconfig("Área para el SPT", state="normal")


# Crear la barra de menú
menu_bar = Menu(root)

# Crear el menú "Archivo"
file_menu = Menu(menu_bar, tearoff=0)
file_menu.add_command(label="Nuevo", command=abrir_nuevo)
file_menu.add_command(label="Abrir", command=abrir_archivo)
file_menu.add_command(label="Guardar")
menu_bar.add_cascade(label="Archivo", menu=file_menu)

# Crear el menú "Edición"
edit_menu = Menu(menu_bar, tearoff=0)
edit_menu.add_command(label="Datos de proyecto", command=abrir_datos_proyecto)
edit_menu.add_command(label="Datos de resistencia", command=abrir_datos_resistencia)
edit_menu.add_command(label="Datos de transformador", command=abrir_datos_transformador)
edit_menu.add_command(label="Datos de conductor", command=abrir_datos_conductor)
edit_menu.add_command(label="Área para el SPT", command=abrir_ventana_spt)
menu_bar.add_cascade(label="Edición", menu=edit_menu)

# Crear el menú "Análisis"
analisis_spt = Menu(menu_bar, tearoff=0)
analisis_spt.add_command(label="Diseño de malla SPT",command=abrir_diseno_malla_spt)
analisis_spt.add_command(label="Análisis de cortocircuito", command=abrir_analisis_cortocircuito)
analisis_spt.add_command(label="Análisis de Resistividad", command=abrir_analisis_resistividad)
menu_bar.add_cascade(label="Análisis", menu=analisis_spt)

# Crear el menú "Reportes"
reports_menu = Menu(menu_bar, tearoff=0)
reports_menu.add_command(label="Configurar reporte")
reports_menu.add_command(label="Exportar PDF")
reports_menu.add_command(label="Exportar CSV")
reports_menu.add_command(label="Exportar Excel")
reports_menu.add_command(label="Exportar Gráfico de malla")
menu_bar.add_cascade(label="Reportes", menu=reports_menu)


# Crear el menú "Ayuda"
help_menu = Menu(menu_bar, tearoff=0)
help_menu.add_command(label="Ayuda")
help_menu.add_command(label="Contactos")
help_menu.add_command(label="Acerca de")
menu_bar.add_cascade(label="Ayuda", menu=help_menu)

# Configurar la barra de menú en la ventana principal
root.config(menu=menu_bar)

# Deshabilita el menú "Edición" al inicio
deshabilitar_menu_edit()

# Ejecutar la aplicación
root.mainloop()